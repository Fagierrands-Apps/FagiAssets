from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from django.urls import reverse
import json
import csv
import qrcode
import io
import base64
from datetime import datetime, timedelta
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from assets.models import (
    Asset, AssetCategory, Manufacturer, AssetModel, 
    Location, Department, MaintenanceRecord, AssetHistory
)
from users.models import UserActivity
from discovery.models import NetworkRange, DiscoveredDevice, ScanJob
from crm.models import Employee, Task, WorkSession, TimeEntry, Communication
from .forms import (
    UserCreateForm, UserEditForm, EmployeeForm, AssetForm,
    AssetCategoryForm, ManufacturerForm, AssetModelForm,
    LocationForm, DepartmentForm, NetworkRangeForm,
    MaintenanceRecordForm, BulkAssetActionForm
)


def is_admin_user(user):
    """Check if user has admin privileges"""
    return user.is_authenticated and (user.is_staff or user.is_superuser)


@login_required
@user_passes_test(is_admin_user)
def admin_dashboard(request):
    """Main admin dashboard with statistics"""
    # Get statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    total_assets = Asset.objects.count()
    active_assets = Asset.objects.filter(status='active').count()
    
    # Asset status breakdown
    asset_status_stats = Asset.objects.values('status').annotate(count=Count('id'))
    
    # Recent activities
    recent_activities = UserActivity.objects.select_related('user').order_by('-timestamp')[:10]
    
    # Recent asset changes
    recent_asset_changes = AssetHistory.objects.select_related('asset', 'user').order_by('-timestamp')[:10]
    
    # Maintenance due soon
    maintenance_due = MaintenanceRecord.objects.filter(
        status__in=['scheduled', 'in_progress'],
        scheduled_date__lte=timezone.now() + timedelta(days=7)
    ).select_related('asset').order_by('scheduled_date')[:5]
    
    # Discovery statistics
    discovered_devices = DiscoveredDevice.objects.count()
    new_devices = DiscoveredDevice.objects.filter(status='new').count()
    
    context = {
        'total_users': total_users,
        'active_users': active_users,
        'total_assets': total_assets,
        'active_assets': active_assets,
        'asset_status_stats': asset_status_stats,
        'recent_activities': recent_activities,
        'recent_asset_changes': recent_asset_changes,
        'maintenance_due': maintenance_due,
        'discovered_devices': discovered_devices,
        'new_devices': new_devices,
    }
    

    return render(request, 'admin_dashboard/dashboard.html', context)


@login_required
@user_passes_test(is_admin_user)
def unified_dashboard(request):
    """Unified dashboard combining all key metrics and information for admin users"""
    # User and Asset Statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    total_assets = Asset.objects.count()
    active_assets = Asset.objects.filter(status='active').count()
    maintenance_assets = Asset.objects.filter(status='maintenance').count()
    
    # Asset status breakdown
    asset_status_stats = Asset.objects.values('status').annotate(count=Count('id'))
    
    # CRM Statistics
    crm_stats = {}
    try:
        from crm.models import Customer, Lead, Employee, Task, WorkSession
        from django.db.models import Sum
        crm_stats = {
            'total_customers': Customer.objects.count(),
            'total_leads': Lead.objects.filter(status='open').count(),
            'active_employees': Employee.objects.filter(employment_status='active').count(),
            'pending_tasks': Task.objects.filter(status='pending').count(),
            'new_customers_this_month': Customer.objects.filter(
                created_at__gte=timezone.now().replace(day=1)
            ).count(),
        }
        
        # Employee time tracking summary
        current_month = timezone.now().replace(day=1)
        total_work_hours = WorkSession.objects.filter(
            date__gte=current_month
        ).aggregate(total=Sum('worked_hours'))['total'] or 0
        
        crm_stats['total_work_hours_this_month'] = round(total_work_hours, 1)
        
    except Exception:
        crm_stats = {
            'total_customers': 0,
            'total_leads': 0,
            'active_employees': 0,
            'pending_tasks': 0,
            'new_customers_this_month': 0,
            'total_work_hours_this_month': 0,
        }
    
    # Recent activities (combined from multiple sources)
    recent_activities = UserActivity.objects.select_related('user').order_by('-timestamp')[:8]
    
    # Recent asset changes
    recent_asset_changes = AssetHistory.objects.select_related('asset', 'user').order_by('-timestamp')[:5]
    
    # Maintenance due soon
    maintenance_due = MaintenanceRecord.objects.filter(
        status__in=['scheduled', 'in_progress'],
        scheduled_date__lte=timezone.now() + timedelta(days=7)
    ).select_related('asset').order_by('scheduled_date')[:5]
    
    # Discovery statistics
    discovered_devices = DiscoveredDevice.objects.count()
    new_devices = DiscoveredDevice.objects.filter(status='new').count()
    
    # Assets by category
    assets_by_category = AssetCategory.objects.annotate(
        asset_count=Count('assetmodel__asset')
    ).order_by('-asset_count')[:6]
    
    # Quick action stats
    quick_stats = {
        'users_created_today': User.objects.filter(
            date_joined__date=timezone.now().date()
        ).count(),
        'assets_added_today': Asset.objects.filter(
            created_at__date=timezone.now().date()
        ).count(),
        'maintenance_completed_today': MaintenanceRecord.objects.filter(
            status='completed',
            completed_date__date=timezone.now().date()
        ).count(),
    }
    
    context = {
        # User and Asset Stats
        'total_users': total_users,
        'active_users': active_users,
        'total_assets': total_assets,
        'active_assets': active_assets,
        'maintenance_assets': maintenance_assets,
        'asset_status_stats': asset_status_stats,
        
        # CRM Stats
        'crm_stats': crm_stats,
        
        # Activity Feeds
        'recent_activities': recent_activities,
        'recent_asset_changes': recent_asset_changes,
        
        # Maintenance & Discovery
        'maintenance_due': maintenance_due,
        'discovered_devices': discovered_devices,
        'new_devices': new_devices,
        
        # Analytics
        'assets_by_category': assets_by_category,
        'quick_stats': quick_stats,
    }
    
    return render(request, 'admin_dashboard/unified_dashboard.html', context)


# User Management Views
@login_required
@user_passes_test(is_admin_user)
def user_list(request):
    """List all users with search and filtering"""
    users = User.objects.select_related('employee_profile').all()

    # Search functionality
    search = request.GET.get('search')
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(employee_profile__employee_id__icontains=search)
        )

    # Filter by status
    status = request.GET.get('status')
    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)
    elif status == 'staff':
        users = users.filter(is_staff=True)

    # Filter by role
    role = request.GET.get('role')
    if role:
        users = users.filter(employee_profile__role=role)

    # Get departments for filter dropdown
    departments = Department.objects.all()

    # Pagination
    paginator = Paginator(users, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'role': role,
        'departments': departments,
    }

    return render(request, 'admin_dashboard/user_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def user_create(request):
    """Create a new user"""
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        employee_form = EmployeeForm(request.POST)
        
        if form.is_valid() and employee_form.is_valid():
            with transaction.atomic():
                user = form.save()
                # Create employee profile with form data
                employee = employee_form.save(commit=False)
                employee.user = user
                employee.save()
                
                # Log activity
                UserActivity.objects.create(
                    user=request.user,
                    action='create',
                    description=f'Created user: {user.username}',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(request, f'User {user.username} created successfully.')
                return redirect('admin_dashboard:user_detail', user_id=user.id)
    else:
        form = UserCreateForm()
        employee_form = EmployeeForm()
    
    # Get role-department mapping for JavaScript
    role_department_mapping = {
        'call_center': 'Customer Service',
        'sales': 'Sales',
        'admin': 'Administration',
        'accountant': 'Finance',
        'receptionist': 'Front Desk',
        'project_support': 'Project Support',
        'hr': 'Human Resources',
        'project_manager': 'Project Management',
        'dev': 'Development',
        'user': 'General',
    }

    context = {
        'form': form,
        'employee_form': employee_form,
        'title': 'Create User',
        'role_department_mapping': json.dumps(role_department_mapping),
    }
    
    return render(request, 'admin_dashboard/user_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def user_detail(request, user_id):
    """View user details"""
    user = get_object_or_404(User, id=user_id)
    
    # Get user's assets
    assigned_assets = Asset.objects.filter(assigned_to=user).select_related('model__manufacturer')
    
    # Get user's recent activities
    recent_activities = UserActivity.objects.filter(user=user).order_by('-timestamp')[:10]
    
    context = {
        'user_obj': user,
        'assigned_assets': assigned_assets,
        'recent_activities': recent_activities,
    }
    
    return render(request, 'admin_dashboard/user_detail.html', context)


@login_required
@user_passes_test(is_admin_user)
def user_edit(request, user_id):
    """Edit user details"""
    user = get_object_or_404(User, id=user_id)
    employee, created = Employee.objects.get_or_create(
        user=user,
        defaults={
            'employee_id': f"EMP{user.id:04d}",
            'position': 'Employee',
            'hire_date': timezone.now().date(),
        }
    )

    def get_department_for_role(role):
        """Get or create department based on role"""
        role_department_mapping = {
            'call_center': 'Customer Service',
            'sales': 'Sales',
            'admin': 'Administration',
            'accountant': 'Finance',
            'receptionist': 'Front Desk',
            'project_support': 'Project Support',
            'hr': 'Human Resources',
            'project_manager': 'Project Management',
            'dev': 'Development',
            'user': 'General',
        }

        department_name = role_department_mapping.get(role, 'General')

        # Get or create the department
        department, created = Department.objects.get_or_create(
            name=department_name,
            defaults={'description': f'Department for {department_name} role'}
        )

        return department

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        employee_form = EmployeeForm(request.POST, instance=employee)

        if form.is_valid() and employee_form.is_valid():
            # Ensure hire_date is set if it's None
            if employee_form.instance.hire_date is None:
                employee_form.instance.hire_date = timezone.now().date()

            # Save the user form first
            form.save()

            # Update role from the user form if provided and assign department
            if 'role' in form.cleaned_data and form.cleaned_data['role']:
                employee.role = form.cleaned_data['role']
                # Auto-assign department based on role
                employee.department = get_department_for_role(employee.role)
                employee.save()

            # Save the employee form
            employee_form.save()
            
            # Sync Employee position and department back to UserProfile
            try:
                if hasattr(user, 'profile'):
                    user.profile.job_title = employee.position
                    user.profile.department = employee.department
                    user.profile.save()
            except Exception as e:
                print(f"Error syncing job title/department for user {user.username}: {e}")

            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='update',
                description=f'Updated user: {user.username}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )

            messages.success(request, f'User {user.username} updated successfully.')
            return redirect('admin_dashboard:user_detail', user_id=user.id)
    else:
        form = UserEditForm(instance=user)
        employee_form = EmployeeForm(instance=employee)

        # Set initial role value from employee profile
        if employee.role:
            form.fields['role'].initial = employee.role

    # Get role-department mapping for JavaScript
    role_department_mapping = {
        'call_center': 'Customer Service',
        'sales': 'Sales',
        'admin': 'Administration',
        'accountant': 'Finance',
        'receptionist': 'Front Desk',
        'project_support': 'Project Support',
        'hr': 'Human Resources',
        'project_manager': 'Project Management',
        'dev': 'Development',
        'user': 'General',
    }

    context = {
        'form': form,
        'employee_form': employee_form,
        'user_obj': user,
        'title': f'Edit User: {user.username}',
        'role_department_mapping': json.dumps(role_department_mapping),
    }

    return render(request, 'admin_dashboard/user_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def user_delete(request, user_id):
    """Delete a user"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        
        # Log activity
        UserActivity.objects.create(
            user=request.user,
            action='delete',
            description=f'Deleted user: {username}',
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        messages.success(request, f'User {username} deleted successfully.')
        return redirect('admin_dashboard:user_list')
    
    context = {
        'user_obj': user,
        'title': f'Delete User: {user.username}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


@login_required
@user_passes_test(is_admin_user)
def user_toggle_active(request, user_id):
    """Toggle user active status"""
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    
    status = 'activated' if user.is_active else 'deactivated'
    messages.success(request, f'User {user.username} {status} successfully.')
    
    # Log activity
    UserActivity.objects.create(
        user=request.user,
        action='update',
        description=f'{status.title()} user: {user.username}',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return redirect('admin_dashboard:user_detail', user_id=user.id)


# Asset Management Views
@login_required
@user_passes_test(is_admin_user)
def asset_admin_list(request):
    """List all assets with admin controls"""
    assets = Asset.objects.select_related('model__manufacturer', 'assigned_to', 'location', 'department').all()
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        assets = assets.filter(
            Q(asset_tag__icontains=search) |
            Q(name__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(model__name__icontains=search) |
            Q(model__manufacturer__name__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        assets = assets.filter(status=status)
    
    # Filter by category
    category = request.GET.get('category')
    if category:
        assets = assets.filter(model__category_id=category)
    
    # Pagination
    paginator = Paginator(assets, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get categories for filter
    categories = AssetCategory.objects.all()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'category': category,
        'categories': categories,
        'status_choices': Asset.STATUS_CHOICES,
    }
    
    return render(request, 'admin_dashboard/asset_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_create(request):
    """Create a new asset"""
    if request.method == 'POST':
        form = AssetForm(request.POST, request.FILES)
        
        if form.is_valid():
            asset = form.save()
            
            # Create history record
            AssetHistory.objects.create(
                asset=asset,
                action='created',
                description=f'Asset created by {request.user.username}',
                user=request.user
            )
            
            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='create',
                description=f'Created asset: {asset.asset_tag}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, f'Asset {asset.asset_tag} created successfully.')
            return redirect('admin_dashboard:asset_detail', asset_id=asset.id)
    else:
        form = AssetForm()
    
    context = {
        'form': form,
        'title': 'Create Asset',
    }
    
    return render(request, 'admin_dashboard/asset_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_admin_detail(request, asset_id):
    """View asset details with admin controls"""
    asset = get_object_or_404(Asset, id=asset_id)
    
    # Get asset history
    history = AssetHistory.objects.filter(asset=asset).select_related('user').order_by('-timestamp')[:10]
    
    # Get maintenance records
    maintenance_records = MaintenanceRecord.objects.filter(asset=asset).select_related('performed_by').order_by('-scheduled_date')[:5]
    
    context = {
        'asset': asset,
        'history': history,
        'maintenance_records': maintenance_records,
    }
    
    return render(request, 'admin_dashboard/asset_detail.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_edit(request, asset_id):
    """Edit asset details"""
    asset = get_object_or_404(Asset, id=asset_id)
    
    if request.method == 'POST':
        form = AssetForm(request.POST, request.FILES, instance=asset)
        
        if form.is_valid():
            # Store previous values for history
            previous_values = {
                'status': asset.status,
                'assigned_to': asset.assigned_to.username if asset.assigned_to else None,
                'location': asset.location.name if asset.location else None,
                'department': asset.department.name if asset.department else None,
            }
            
            asset = form.save()
            
            # Store new values
            new_values = {
                'status': asset.status,
                'assigned_to': asset.assigned_to.username if asset.assigned_to else None,
                'location': asset.location.name if asset.location else None,
                'department': asset.department.name if asset.department else None,
            }
            
            # Create history record
            AssetHistory.objects.create(
                asset=asset,
                action='updated',
                description=f'Asset updated by {request.user.username}',
                user=request.user,
                previous_values=previous_values,
                new_values=new_values
            )
            
            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='update',
                description=f'Updated asset: {asset.asset_tag}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, f'Asset {asset.asset_tag} updated successfully.')
            return redirect('admin_dashboard:asset_detail', asset_id=asset.id)
    else:
        form = AssetForm(instance=asset)
    
    context = {
        'form': form,
        'asset': asset,
        'title': f'Edit Asset: {asset.asset_tag}',
    }
    
    return render(request, 'admin_dashboard/asset_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_delete(request, asset_id):
    """Delete an asset"""
    asset = get_object_or_404(Asset, id=asset_id)
    
    if request.method == 'POST':
        asset_tag = asset.asset_tag
        asset.delete()
        
        # Log activity
        UserActivity.objects.create(
            user=request.user,
            action='delete',
            description=f'Deleted asset: {asset_tag}',
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        messages.success(request, f'Asset {asset_tag} deleted successfully.')
        return redirect('admin_dashboard:asset_list')
    
    context = {
        'asset': asset,
        'title': f'Delete Asset: {asset.asset_tag}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# Category Management Views
@login_required
@user_passes_test(is_admin_user)
def category_list(request):
    """List all asset categories"""
    categories = AssetCategory.objects.select_related('parent').annotate(
        asset_count=Count('assetmodel__asset')
    ).order_by('name')
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'admin_dashboard/category_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def category_create(request):
    """Create a new category"""
    if request.method == 'POST':
        form = AssetCategoryForm(request.POST)
        
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category {category.name} created successfully.')
            return redirect('admin_dashboard:category_list')
    else:
        form = AssetCategoryForm()
    
    context = {
        'form': form,
        'title': 'Create Category',
    }
    
    return render(request, 'admin_dashboard/category_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def category_edit(request, category_id):
    """Edit a category"""
    category = get_object_or_404(AssetCategory, id=category_id)
    
    if request.method == 'POST':
        form = AssetCategoryForm(request.POST, instance=category)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Category {category.name} updated successfully.')
            return redirect('admin_dashboard:category_list')
    else:
        form = AssetCategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
        'title': f'Edit Category: {category.name}',
    }
    
    return render(request, 'admin_dashboard/category_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def category_delete(request, category_id):
    """Delete a category"""
    category = get_object_or_404(AssetCategory, id=category_id)
    
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Category {category_name} deleted successfully.')
        return redirect('admin_dashboard:category_list')
    
    context = {
        'category': category,
        'title': f'Delete Category: {category.name}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# Manufacturer Management Views
@login_required
@user_passes_test(is_admin_user)
def manufacturer_list(request):
    """List all manufacturers"""
    manufacturers = Manufacturer.objects.annotate(
        model_count=Count('assetmodel'),
        asset_count=Count('assetmodel__asset')
    ).order_by('name')
    
    context = {
        'manufacturers': manufacturers,
    }
    
    return render(request, 'admin_dashboard/manufacturer_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def manufacturer_create(request):
    """Create a new manufacturer"""
    if request.method == 'POST':
        form = ManufacturerForm(request.POST)
        
        if form.is_valid():
            manufacturer = form.save()
            messages.success(request, f'Manufacturer {manufacturer.name} created successfully.')
            return redirect('admin_dashboard:manufacturer_list')
    else:
        form = ManufacturerForm()
    
    context = {
        'form': form,
        'title': 'Create Manufacturer',
    }
    
    return render(request, 'admin_dashboard/manufacturer_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def manufacturer_edit(request, manufacturer_id):
    """Edit a manufacturer"""
    manufacturer = get_object_or_404(Manufacturer, id=manufacturer_id)
    
    if request.method == 'POST':
        form = ManufacturerForm(request.POST, instance=manufacturer)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Manufacturer {manufacturer.name} updated successfully.')
            return redirect('admin_dashboard:manufacturer_list')
    else:
        form = ManufacturerForm(instance=manufacturer)
    
    context = {
        'form': form,
        'manufacturer': manufacturer,
        'title': f'Edit Manufacturer: {manufacturer.name}',
    }
    
    return render(request, 'admin_dashboard/manufacturer_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def manufacturer_delete(request, manufacturer_id):
    """Delete a manufacturer"""
    manufacturer = get_object_or_404(Manufacturer, id=manufacturer_id)
    
    if request.method == 'POST':
        manufacturer_name = manufacturer.name
        manufacturer.delete()
        messages.success(request, f'Manufacturer {manufacturer_name} deleted successfully.')
        return redirect('admin_dashboard:manufacturer_list')
    
    context = {
        'manufacturer': manufacturer,
        'title': f'Delete Manufacturer: {manufacturer.name}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# Asset Model Management Views
@login_required
@user_passes_test(is_admin_user)
def asset_model_list(request):
    """List all asset models"""
    models = AssetModel.objects.select_related('manufacturer', 'category').annotate(
        asset_count=Count('asset')
    ).order_by('manufacturer__name', 'name')
    
    context = {
        'models': models,
    }
    
    return render(request, 'admin_dashboard/asset_model_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_model_create(request):
    """Create a new asset model"""
    if request.method == 'POST':
        form = AssetModelForm(request.POST, request.FILES)
        
        if form.is_valid():
            model = form.save()
            messages.success(request, f'Asset model {model.name} created successfully.')
            return redirect('admin_dashboard:asset_model_list')
    else:
        form = AssetModelForm()
    
    context = {
        'form': form,
        'title': 'Create Asset Model',
    }
    
    return render(request, 'admin_dashboard/asset_model_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_model_edit(request, model_id):
    """Edit an asset model"""
    model = get_object_or_404(AssetModel, id=model_id)
    
    if request.method == 'POST':
        form = AssetModelForm(request.POST, request.FILES, instance=model)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Asset model {model.name} updated successfully.')
            return redirect('admin_dashboard:asset_model_list')
    else:
        form = AssetModelForm(instance=model)
    
    context = {
        'form': form,
        'model': model,
        'title': f'Edit Asset Model: {model.name}',
    }
    
    return render(request, 'admin_dashboard/asset_model_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def asset_model_delete(request, model_id):
    """Delete an asset model"""
    model = get_object_or_404(AssetModel, id=model_id)
    
    if request.method == 'POST':
        model_name = model.name
        model.delete()
        messages.success(request, f'Asset model {model_name} deleted successfully.')
        return redirect('admin_dashboard:asset_model_list')
    
    context = {
        'model': model,
        'title': f'Delete Asset Model: {model.name}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# Location Management Views
@login_required
@user_passes_test(is_admin_user)
def location_list(request):
    """List all locations"""
    locations = Location.objects.annotate(
        asset_count=Count('asset'),
        user_count=Count('userprofile')
    ).order_by('name')
    
    context = {
        'locations': locations,
    }
    
    return render(request, 'admin_dashboard/location_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def location_create(request):
    """Create a new location"""
    if request.method == 'POST':
        form = LocationForm(request.POST)
        
        if form.is_valid():
            location = form.save()
            messages.success(request, f'Location {location.name} created successfully.')
            return redirect('admin_dashboard:location_list')
    else:
        form = LocationForm()
    
    context = {
        'form': form,
        'title': 'Create Location',
    }
    
    return render(request, 'admin_dashboard/location_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def location_edit(request, location_id):
    """Edit a location"""
    location = get_object_or_404(Location, id=location_id)
    
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Location {location.name} updated successfully.')
            return redirect('admin_dashboard:location_list')
    else:
        form = LocationForm(instance=location)
    
    context = {
        'form': form,
        'location': location,
        'title': f'Edit Location: {location.name}',
    }
    
    return render(request, 'admin_dashboard/location_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def location_delete(request, location_id):
    """Delete a location"""
    location = get_object_or_404(Location, id=location_id)
    
    if request.method == 'POST':
        location_name = location.name
        location.delete()
        messages.success(request, f'Location {location_name} deleted successfully.')
        return redirect('admin_dashboard:location_list')
    
    context = {
        'location': location,
        'title': f'Delete Location: {location.name}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# Department Management Views
@login_required
@user_passes_test(is_admin_user)
def department_list(request):
    """List all departments"""
    departments = Department.objects.select_related('manager').annotate(
        asset_count=Count('asset'),
        user_count=Count('userprofile')
    ).order_by('name')
    
    context = {
        'departments': departments,
    }
    
    return render(request, 'admin_dashboard/department_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def department_create(request):
    """Create a new department"""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Department {department.name} created successfully.')
            return redirect('admin_dashboard:department_list')
    else:
        form = DepartmentForm()
    
    context = {
        'form': form,
        'title': 'Create Department',
    }
    
    return render(request, 'admin_dashboard/department_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def department_edit(request, department_id):
    """Edit a department"""
    department = get_object_or_404(Department, id=department_id)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Department {department.name} updated successfully.')
            return redirect('admin_dashboard:department_list')
    else:
        form = DepartmentForm(instance=department)
    
    context = {
        'form': form,
        'department': department,
        'title': f'Edit Department: {department.name}',
    }
    
    return render(request, 'admin_dashboard/department_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def department_delete(request, department_id):
    """Delete a department"""
    department = get_object_or_404(Department, id=department_id)
    
    if request.method == 'POST':
        department_name = department.name
        department.delete()
        messages.success(request, f'Department {department_name} deleted successfully.')
        return redirect('admin_dashboard:department_list')
    
    context = {
        'department': department,
        'title': f'Delete Department: {department.name}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# Discovery Management Views
@login_required
@user_passes_test(is_admin_user)
def discovery_admin(request):
    """Discovery administration dashboard"""
    # Get statistics
    total_ranges = NetworkRange.objects.count()
    active_ranges = NetworkRange.objects.filter(is_active=True).count()
    discovered_devices = DiscoveredDevice.objects.count()
    new_devices = DiscoveredDevice.objects.filter(status='new').count()
    
    # Recent scan jobs
    recent_scans = ScanJob.objects.select_related('created_by').order_by('-created_at')[:5]
    
    # Recent discovered devices
    recent_devices = DiscoveredDevice.objects.select_related('network_range').order_by('-last_seen')[:10]
    
    context = {
        'total_ranges': total_ranges,
        'active_ranges': active_ranges,
        'discovered_devices': discovered_devices,
        'new_devices': new_devices,
        'recent_scans': recent_scans,
        'recent_devices': recent_devices,
    }
    
    return render(request, 'admin_dashboard/discovery_admin.html', context)


@login_required
@user_passes_test(is_admin_user)
def network_range_list(request):
    """List all network ranges"""
    ranges = NetworkRange.objects.annotate(
        device_count=Count('discovered_devices')
    ).order_by('name')
    
    context = {
        'ranges': ranges,
    }
    
    return render(request, 'admin_dashboard/network_range_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def network_range_create(request):
    """Create a new network range"""
    if request.method == 'POST':
        form = NetworkRangeForm(request.POST)
        
        if form.is_valid():
            network_range = form.save()
            messages.success(request, f'Network range {network_range.name} created successfully.')
            return redirect('admin_dashboard:network_range_list')
    else:
        form = NetworkRangeForm()
    
    context = {
        'form': form,
        'title': 'Create Network Range',
    }
    
    return render(request, 'admin_dashboard/network_range_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def network_range_edit(request, range_id):
    """Edit a network range"""
    network_range = get_object_or_404(NetworkRange, id=range_id)
    
    if request.method == 'POST':
        form = NetworkRangeForm(request.POST, instance=network_range)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Network range {network_range.name} updated successfully.')
            return redirect('admin_dashboard:network_range_list')
    else:
        form = NetworkRangeForm(instance=network_range)
    
    context = {
        'form': form,
        'network_range': network_range,
        'title': f'Edit Network Range: {network_range.name}',
    }
    
    return render(request, 'admin_dashboard/network_range_form.html', context)


@login_required
@user_passes_test(is_admin_user)
def network_range_delete(request, range_id):
    """Delete a network range"""
    network_range = get_object_or_404(NetworkRange, id=range_id)
    
    if request.method == 'POST':
        range_name = network_range.name
        network_range.delete()
        messages.success(request, f'Network range {range_name} deleted successfully.')
        return redirect('admin_dashboard:network_range_list')
    
    context = {
        'network_range': network_range,
        'title': f'Delete Network Range: {network_range.name}',
    }
    
    return render(request, 'admin_dashboard/confirm_delete.html', context)


# System Settings Views
@login_required
@user_passes_test(is_admin_user)
def system_settings(request):
    """System settings and configuration"""
    context = {
        'title': 'System Settings',
    }
    
    return render(request, 'admin_dashboard/system_settings.html', context)


@login_required
@user_passes_test(is_admin_user)
def system_backup(request):
    """System backup functionality"""
    if request.method == 'POST':
        # Create backup logic here
        messages.success(request, 'System backup created successfully.')
    
    context = {
        'title': 'System Backup',
    }
    
    return render(request, 'admin_dashboard/system_backup.html', context)


@login_required
@user_passes_test(is_admin_user)
def system_logs(request):
    """View system logs"""
    activities = UserActivity.objects.select_related('user').order_by('-timestamp')
    
    # Pagination
    paginator = Paginator(activities, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'System Logs',
    }
    
    return render(request, 'admin_dashboard/system_logs.html', context)


# AJAX Views
@login_required
@user_passes_test(is_admin_user)
def ajax_user_search(request):
    """AJAX endpoint for user search"""
    query = request.GET.get('q', '')
    users = User.objects.filter(
        Q(username__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query)
    )[:10]
    
    results = [
        {
            'id': user.id,
            'text': f"{user.get_full_name() or user.username} ({user.username})"
        }
        for user in users
    ]
    
    return JsonResponse({'results': results})


@login_required
@user_passes_test(is_admin_user)
def ajax_asset_search(request):
    """AJAX endpoint for asset search"""
    query = request.GET.get('q', '')
    assets = Asset.objects.filter(
        Q(asset_tag__icontains=query) |
        Q(name__icontains=query) |
        Q(serial_number__icontains=query)
    ).select_related('model__manufacturer')[:10]
    
    results = [
        {
            'id': asset.id,
            'text': f"{asset.asset_tag} - {asset.name} ({asset.model.manufacturer.name} {asset.model.name})"
        }
        for asset in assets
    ]
    
    return JsonResponse({'results': results})


@login_required
@user_passes_test(is_admin_user)
def user_qr_code(request, user_id):
    """Generate QR code for user details"""
    user = get_object_or_404(User, id=user_id)
    
    # Create public user URL for QR scanning (no login required)
    from django.conf import settings
    if getattr(settings, 'QR_CODE_BASE_URL', ''):
        user_details_url = f"{settings.QR_CODE_BASE_URL.rstrip('/')}/users/{user.id}/public/"
    else:
        user_details_url = request.build_absolute_uri(
            reverse('users:user_public_view', args=[user.id])
        )
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(user_details_url)
    qr.make(fit=True)
    
    # Create QR code image
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    # Create a larger image with user info
    img_width = 400
    img_height = 500
    img = Image.new('RGB', (img_width, img_height), 'white')
    
    # Paste QR code
    qr_img = qr_img.resize((300, 300))
    img.paste(qr_img, (50, 50))
    
    # Add text information
    draw = ImageDraw.Draw(img)
    
    try:
        # Try to use a better font
        font_large = ImageFont.truetype("arial.ttf", 20)
        font_medium = ImageFont.truetype("arial.ttf", 16)
        font_small = ImageFont.truetype("arial.ttf", 12)
    except:
        # Fallback to default font
        font_large = ImageFont.load_default()
        font_medium = ImageFont.load_default()
        font_small = ImageFont.load_default()
    
    # Add user information
    y_pos = 370
    
    # User name
    name = user.get_full_name() or user.username
    draw.text((50, y_pos), f"Name: {name}", fill="black", font=font_medium)
    y_pos += 25
    
    # Username
    draw.text((50, y_pos), f"Username: @{user.username}", fill="black", font=font_small)
    y_pos += 20
    
    # Employee ID if available
    if hasattr(user, 'profile') and user.profile.employee_id:
        draw.text((50, y_pos), f"Employee ID: {user.profile.employee_id}", fill="black", font=font_small)
        y_pos += 20
    
    # Department if available
    if hasattr(user, 'profile') and user.profile.department:
        draw.text((50, y_pos), f"Department: {user.profile.department.name}", fill="black", font=font_small)
        y_pos += 20
    
    # Add timestamp
    draw.text((50, y_pos), f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", 
              fill="gray", font=font_small)
    
    # Convert to HTTP response
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="user_{user.username}_qr.png"'
    
    return response


@login_required
@user_passes_test(is_admin_user)
def user_qr_details(request, user_id):
    """Display user details when QR code is scanned"""
    user = get_object_or_404(User, id=user_id)
    
    # Get user profile and related information
    try:
        profile = user.profile
    except:
        profile = None
    
    # Get user's assigned assets
    assigned_assets = Asset.objects.filter(assigned_to=user).select_related(
        'model__manufacturer', 'model__category', 'location'
    )
    
    # Get recent activities
    recent_activities = UserActivity.objects.filter(user=user).order_by('-timestamp')[:10]
    
    context = {
        'user': user,
        'profile': profile,
        'assigned_assets': assigned_assets,
        'recent_activities': recent_activities,
        'is_qr_view': True,
    }
    
    return render(request, 'admin_dashboard/user_qr_details.html', context)


@login_required
@user_passes_test(is_admin_user)
def bulk_user_qr_codes(request):
    """Generate QR codes for multiple users"""
    if request.method == 'POST':
        user_ids = request.POST.getlist('selected_users')
        if not user_ids:
            messages.error(request, 'No users selected.')
            return redirect('admin_dashboard:user_list')
        
        users = User.objects.filter(id__in=user_ids)
        
        # Create a ZIP file with all QR codes
        import zipfile
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="user_qr_codes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip"'
        
        with zipfile.ZipFile(response, 'w') as zip_file:
            for user in users:
                # Generate QR code for each user (public URL, no login required)
                from django.conf import settings
                if getattr(settings, 'QR_CODE_BASE_URL', ''):
                    user_details_url = f"{settings.QR_CODE_BASE_URL.rstrip('/')}/users/{user.id}/public/"
                else:
                    user_details_url = request.build_absolute_uri(
                        reverse('users:user_public_view', args=[user.id])
                    )
                
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(user_details_url)
                qr.make(fit=True)
                
                qr_img = qr.make_image(fill_color="black", back_color="white")
                
                # Create enhanced image with user info
                img_width = 400
                img_height = 500
                img = Image.new('RGB', (img_width, img_height), 'white')
                
                qr_img = qr_img.resize((300, 300))
                img.paste(qr_img, (50, 50))
                
                draw = ImageDraw.Draw(img)
                
                try:
                    font_medium = ImageFont.truetype("arial.ttf", 16)
                    font_small = ImageFont.truetype("arial.ttf", 12)
                except:
                    font_medium = ImageFont.load_default()
                    font_small = ImageFont.load_default()
                
                y_pos = 370
                name = user.get_full_name() or user.username
                draw.text((50, y_pos), f"Name: {name}", fill="black", font=font_medium)
                y_pos += 25
                
                draw.text((50, y_pos), f"Username: @{user.username}", fill="black", font=font_small)
                y_pos += 20
                
                if hasattr(user, 'profile') and user.profile.employee_id:
                    draw.text((50, y_pos), f"Employee ID: {user.profile.employee_id}", fill="black", font=font_small)
                    y_pos += 20
                
                draw.text((50, y_pos), f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", 
                          fill="gray", font=font_small)
                
                # Save to buffer and add to ZIP
                buffer = io.BytesIO()
                img.save(buffer, format='PNG')
                buffer.seek(0)
                
                filename = f"user_{user.username}_qr.png"
                zip_file.writestr(filename, buffer.getvalue())
        
        return response
    
    # GET request - show user selection form
    users = User.objects.all().order_by('username')
    return render(request, 'admin_dashboard/bulk_user_qr.html', {'users': users})


# =============================================================================
# EMPLOYEE MANAGEMENT VIEWS
# =============================================================================

@login_required
@user_passes_test(is_admin_user)
def employee_list(request):
    """List all employees with search and filtering"""
    search = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    
    employees = Employee.objects.select_related('user', 'department', 'manager').all()
    
    if search:
        employees = employees.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(employee_id__icontains=search) |
            Q(position__icontains=search)
        )
    
    if department_filter:
        employees = employees.filter(department_id=department_filter)
    
    if status_filter:
        employees = employees.filter(employment_status=status_filter)
    
    # Pagination
    paginator = Paginator(employees, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get departments for filter
    from crm.models import Department
    departments = Department.objects.all()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'departments': departments,
        'status_choices': Employee.EMPLOYMENT_STATUS_CHOICES,
    }
    
    return render(request, 'admin_dashboard/employee_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def employee_detail(request, employee_id):
    """Employee detail view with tasks and time tracking"""
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Get recent tasks
    recent_tasks = employee.assigned_tasks.all()[:10]
    
    # Get recent time entries
    recent_time_entries = employee.time_entries.all()[:10]
    
    # Get current month work sessions
    current_month = timezone.now().replace(day=1)
    work_sessions = employee.work_sessions.filter(
        date__gte=current_month
    ).order_by('-date')
    
    # Calculate total hours this month
    total_hours = sum([ws.worked_hours or 0 for ws in work_sessions])
    
    context = {
        'employee': employee,
        'recent_tasks': recent_tasks,
        'recent_time_entries': recent_time_entries,
        'work_sessions': work_sessions,
        'total_hours': total_hours,
    }
    
    return render(request, 'admin_dashboard/employee_detail.html', context)


@login_required
@user_passes_test(is_admin_user)
def task_assign(request):
    """Assign tasks to employees"""
    if request.method == 'POST':
        try:
            task = Task.objects.create(
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                assigned_to_id=request.POST.get('assigned_to'),
                assigned_by=request.user,
                priority=request.POST.get('priority', 'medium'),
                due_date=request.POST.get('due_date') if request.POST.get('due_date') else None,
                estimated_hours=request.POST.get('estimated_hours') if request.POST.get('estimated_hours') else None,
                customer_id=request.POST.get('customer') if request.POST.get('customer') else None,
                lead_id=request.POST.get('lead') if request.POST.get('lead') else None,
                asset_id=request.POST.get('asset') if request.POST.get('asset') else None,
            )
            
            messages.success(request, f'Task "{task.title}" assigned to {task.assigned_to.full_name}')
            return redirect('/admin-dashboard/tasks/')
            
        except Exception as e:
            messages.error(request, f'Error creating task: {str(e)}')
    
    # Get data for form
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    from crm.models import Customer, Lead
    customers = Customer.objects.filter(status='active').order_by('first_name')
    leads = Lead.objects.exclude(status__in=['won', 'lost']).order_by('first_name')
    assets = Asset.objects.filter(status='active').order_by('name')
    
    context = {
        'employees': employees,
        'customers': customers,
        'leads': leads,
        'assets': assets,
        'priority_choices': Task.PRIORITY_CHOICES,
    }
    
    return render(request, 'admin_dashboard/task_assign.html', context)


@login_required
@user_passes_test(is_admin_user)
def task_list(request):
    """List all tasks with filtering"""
    search = request.GET.get('search', '')
    employee_filter = request.GET.get('employee', '')
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    
    tasks = Task.objects.select_related('assigned_to__user', 'assigned_by').all()
    
    if search:
        tasks = tasks.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search)
        )
    
    if employee_filter:
        tasks = tasks.filter(assigned_to_id=employee_filter)
    
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)
    
    # Pagination
    paginator = Paginator(tasks, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get employees for filter
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'employee_filter': employee_filter,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'employees': employees,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
    }
    
    return render(request, 'admin_dashboard/task_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def time_tracking_overview(request):
    """Overview of employee time tracking"""
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_filter = request.GET.get('employee', '')
    
    if not start_date:
        start_date = timezone.now().replace(day=1).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get work sessions
    work_sessions = WorkSession.objects.select_related('employee__user').filter(
        date__range=[start_date, end_date]
    )
    
    if employee_filter:
        work_sessions = work_sessions.filter(employee_id=employee_filter)
    
    work_sessions = work_sessions.order_by('-date', 'employee__user__first_name')
    
    # Calculate summary statistics
    total_hours = sum([ws.worked_hours or 0 for ws in work_sessions])
    total_sessions = work_sessions.count()
    avg_hours_per_session = (total_hours / total_sessions) if total_sessions else 0
    
    # Get employees for filter
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    
    # Pagination
    paginator = Paginator(work_sessions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'employee_filter': employee_filter,
        'employees': employees,
        'total_hours': total_hours,
        'total_sessions': total_sessions,
        'avg_hours_per_session': avg_hours_per_session,
    }
    
    return render(request, 'admin_dashboard/time_tracking_overview.html', context)


@login_required
@user_passes_test(is_admin_user)
def download_time_tracking_report(request):
    """Download time tracking reports in CSV, PDF, or Excel format"""
    # Get parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    employee_filter = request.GET.get('employee', '')
    format_type = request.GET.get('format', 'csv')  # csv, pdf, or xlsx
    report_type = request.GET.get('type', 'detailed')  # detailed or summary
    
    # Parse dates (default to current month)
    if not start_date:
        start_date = timezone.now().replace(day=1).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get work sessions
    work_sessions = WorkSession.objects.select_related('employee__user', 'employee__department').filter(
        date__range=[start_date, end_date]
    )
    
    if employee_filter:
        work_sessions = work_sessions.filter(employee_id=employee_filter)
    
    work_sessions = work_sessions.order_by('date', 'employee__user__first_name')
    
    # Generate filename
    filename_date = f"{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
    
    # Route to appropriate generator
    if report_type == 'summary':
        if format_type == 'csv':
            return _generate_summary_csv_report(work_sessions, start_date, end_date, filename_date)
        elif format_type == 'pdf':
            return _generate_summary_pdf_report(work_sessions, start_date, end_date, filename_date)
        elif format_type == 'xlsx':
            return _generate_summary_excel_report(work_sessions, start_date, end_date, filename_date)
    else:  # detailed
        if format_type == 'csv':
            return _generate_csv_report(work_sessions, start_date, end_date, filename_date)
        elif format_type == 'pdf':
            return _generate_pdf_report(work_sessions, start_date, end_date, filename_date)
        elif format_type == 'xlsx':
            return _generate_excel_report(work_sessions, start_date, end_date, filename_date)
    
    return _generate_csv_report(work_sessions, start_date, end_date, filename_date)


def _generate_csv_report(work_sessions, start_date, end_date, filename_date):
    """Generate CSV report"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="TimeTracking_{filename_date}.csv"'
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow([
        'Employee Name', 'Employee ID', 'Department', 'Date', 
        'Punch In', 'Punch Out', 'Break Hours', 'Hours Worked', 'Status'
    ])
    
    # Data rows
    for session in work_sessions:
        writer.writerow([
            session.employee.user.get_full_name() or session.employee.user.username,
            session.employee.employee_id,
            session.employee.department.name if session.employee.department else 'N/A',
            session.date.strftime('%Y-%m-%d'),
            session.punch_in.strftime('%H:%M:%S') if session.punch_in else 'N/A',
            session.punch_out.strftime('%H:%M:%S') if session.punch_out else 'N/A',
            float(session.break_hours) or 0,
            float(session.worked_hours) or 0,
            'Completed' if session.is_complete else 'In Progress',
        ])
    
    # Summary rows
    total_hours = sum([float(ws.worked_hours) or 0 for ws in work_sessions])
    writer.writerow([])
    writer.writerow(['Summary Report'])
    writer.writerow(['Start Date', start_date.strftime('%Y-%m-%d')])
    writer.writerow(['End Date', end_date.strftime('%Y-%m-%d')])
    writer.writerow(['Total Sessions', work_sessions.count()])
    writer.writerow(['Total Hours Worked', f'{total_hours:.2f}'])
    writer.writerow(['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    return response


def _generate_excel_report(work_sessions, start_date, end_date, filename_date):
    """Generate Excel report"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="TimeTracking_{filename_date}.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Time Tracking'
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    summary_font = Font(bold=True, size=10)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 18
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 18
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    
    # Header row
    headers = [
        'Employee Name', 'Employee ID', 'Department', 'Date',
        'Punch In', 'Punch Out', 'Break Hours', 'Hours Worked', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    current_row = 2
    for session in work_sessions:
        cells = [
            session.employee.user.get_full_name() or session.employee.user.username,
            session.employee.employee_id,
            session.employee.department.name if session.employee.department else 'N/A',
            session.date.strftime('%Y-%m-%d'),
            session.punch_in.strftime('%H:%M:%S') if session.punch_in else 'N/A',
            session.punch_out.strftime('%H:%M:%S') if session.punch_out else 'N/A',
            float(session.break_hours) or 0,
            float(session.worked_hours) or 0,
            'Completed' if session.is_complete else 'In Progress',
        ]
        
        for col, value in enumerate(cells, 1):
            cell = worksheet.cell(row=current_row, column=col)
            cell.value = value
            cell.border = border
            if col in [4, 8]:  # Date and Hours - center alignment
                cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
    
    # Summary section
    summary_row = current_row + 2
    total_hours = sum([float(ws.worked_hours) or 0 for ws in work_sessions])
    
    summary_data = [
        ['SUMMARY REPORT', ''],
        ['Start Date', start_date.strftime('%Y-%m-%d')],
        ['End Date', end_date.strftime('%Y-%m-%d')],
        ['Total Sessions', work_sessions.count()],
        ['Total Hours Worked', f'{total_hours:.2f}'],
        ['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for idx, row_data in enumerate(summary_data, 0):
        for col, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=summary_row + idx, column=col)
            cell.value = value
            if idx == 0:  # Title row
                cell.font = summary_font
                cell.fill = summary_fill
            cell.border = border
    
    workbook.save(response)
    return response


def _generate_pdf_report(work_sessions, start_date, end_date, filename_date):
    """Generate PDF report"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="TimeTracking_{filename_date}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1  # Center
    )
    story.append(Paragraph('Time Tracking Report', title_style))
    
    # Period info
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=12
    )
    period_text = f"Report Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    story.append(Paragraph(period_text, info_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [[
        'Employee Name', 'Employee ID', 'Department', 'Date',
        'Punch In', 'Punch Out', 'Break\n(hrs)', 'Hours', 'Status'
    ]]
    
    for session in work_sessions:
        table_data.append([
            session.employee.user.get_full_name() or session.employee.user.username,
            session.employee.employee_id,
            session.employee.department.name if session.employee.department else 'N/A',
            session.date.strftime('%Y-%m-%d'),
            session.punch_in.strftime('%H:%M') if session.punch_in else 'N/A',
            session.punch_out.strftime('%H:%M') if session.punch_out else 'N/A',
            f"{float(session.break_hours) or 0:.2f}",
            f"{float(session.worked_hours) or 0:.1f}",
            'Completed' if session.is_complete else 'In Progress',
        ])
    
    # Create table
    table = Table(table_data, colWidths=[1.4*inch, 0.9*inch, 1.2*inch, 0.9*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.9*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Summary section
    total_hours = sum([float(ws.worked_hours) or 0 for ws in work_sessions])
    summary_text = f"""
    <b>Summary:</b><br/>
    Total Sessions: {work_sessions.count()}<br/>
    Total Hours Worked: {total_hours:.2f}<br/>
    Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    return response


# ==================== SUMMARY REPORT FUNCTIONS ====================

def _calculate_required_work_hours(start_date, end_date):
    """Calculate total required work hours for the date range.
    
    Work schedule: Monday to Saturday, 8 hours per day (8 AM - 5 PM with 1 hour break)
    
    Args:
        start_date: Start date (datetime.date)
        end_date: End date (datetime.date)
    
    Returns:
        Total required hours for the period
    """
    from datetime import timedelta
    
    current = start_date
    total_hours = 0
    
    while current <= end_date:
        # Monday = 0, Sunday = 6
        if current.weekday() < 6:  # Monday (0) to Saturday (5), exclude Sunday (6)
            total_hours += 8  # 8 hours per day
        current += timedelta(days=1)
    
    return total_hours


def _get_summary_data(work_sessions, max_incomplete_age_hours=24, start_date=None, end_date=None):
    """Aggregate work sessions by employee and return summary data sorted by total hours
    
    Includes:
    - Completed sessions with calculated hours
    - Incomplete sessions with real-time calculation (current time - punch_in - breaks)
    - Required work hours for the period (Mon-Sat, 8 hrs/day)
    - Variance calculation (actual vs required)
    
    Args:
        work_sessions: QuerySet of work sessions to aggregate
        max_incomplete_age_hours: Maximum age (in hours) for an incomplete session to be included.
                                 Sessions older than this are excluded (likely abandoned).
                                 Default: 24 hours (1 day)
        start_date: Start date for required hours calculation (datetime.date)
        end_date: End date for required hours calculation (datetime.date)
    """
    from django.db.models import Sum, Count, Q
    from decimal import Decimal
    
    # Calculate total required work hours for the period
    total_required_hours = 0
    if start_date and end_date:
        total_required_hours = _calculate_required_work_hours(start_date, end_date)
    
    summary_dict = {}
    
    for session in work_sessions:
        emp_id = session.employee_id
        if emp_id not in summary_dict:
            summary_dict[emp_id] = {
                'employee_name': session.employee.user.get_full_name() or session.employee.user.username,
                'employee_id': session.employee.employee_id,
                'department': session.employee.department.name if session.employee.department else 'N/A',
                'total_hours': 0,
                'total_productive_hours': 0,
                'total_idle_hours': 0,
                'days_worked': set(),
                'sessions_count': 0,
                'incomplete_sessions': 0,
                'stale_sessions': 0,  # Track excluded stale sessions
            }
        
        # Calculate hours for this session
        if session.is_complete or session.punch_out:
            # Completed session - use calculated hours
            session_hours = float(session.worked_hours or 0)
            productive = float(session.productive_hours or 0)
            idle = float(session.idle_hours or 0)
            
            # Ensure productive + idle = session_hours (fix any data inconsistencies)
            total_productive_idle = productive + idle
            if total_productive_idle > 0 and abs(total_productive_idle - session_hours) > 0.01:
                # If there's a discrepancy, normalize proportionally
                if total_productive_idle > session_hours:
                    # Scale down if productive + idle exceeds total
                    scale_factor = session_hours / total_productive_idle if total_productive_idle > 0 else 0
                    productive = productive * scale_factor
                    idle = idle * scale_factor
                else:
                    # If productive + idle is less than total, distribute the difference proportionally
                    # or assign remainder to idle (assuming unaccounted time is idle)
                    idle = idle + (session_hours - total_productive_idle)
        else:
            # Incomplete session - calculate real-time hours
            if session.punch_in:
                current_time = timezone.now()
                time_elapsed = current_time - session.punch_in
                hours_elapsed = time_elapsed.total_seconds() / 3600
                
                # CHECK: Exclude stale incomplete sessions (likely abandoned)
                if hours_elapsed > max_incomplete_age_hours:
                    # Session is too old and incomplete - skip it
                    summary_dict[emp_id]['stale_sessions'] += 1
                    continue
                
                total_seconds = time_elapsed.total_seconds()
                session_hours = total_seconds / 3600
                break_hours_float = float(session.break_hours or 0)
                session_hours = max(0, session_hours - break_hours_float)
                
                # For incomplete sessions, use stored values if available, otherwise estimate
                productive_stored = float(session.productive_hours or 0)
                idle_stored = float(session.idle_hours or 0)
                
                if productive_stored + idle_stored > 0:
                    # Use stored values and scale to match current session_hours
                    total_stored = productive_stored + idle_stored
                    if total_stored > 0:
                        scale_factor = session_hours / total_stored if total_stored <= session_hours else 1.0
                        productive = productive_stored * scale_factor
                        idle = idle_stored * scale_factor
                        # Ensure they sum to session_hours
                        total_current = productive + idle
                        if total_current > 0:
                            productive = (productive / total_current) * session_hours
                            idle = (idle / total_current) * session_hours
                        else:
                            productive = session_hours
                            idle = 0
                    else:
                        productive = session_hours
                        idle = 0
                else:
                    # No stored values - assume all working time is productive until proven otherwise
                    productive = session_hours
                    idle = 0
            else:
                session_hours = 0
                productive = 0
                idle = 0
            
            summary_dict[emp_id]['incomplete_sessions'] += 1
        
        # Ensure productive + idle equals session_hours (final validation)
        total_productive_idle = productive + idle
        if abs(total_productive_idle - session_hours) > 0.01:
            # Normalize to ensure they sum correctly
            if total_productive_idle > 0:
                scale_factor = session_hours / total_productive_idle
                productive = productive * scale_factor
                idle = idle * scale_factor
            else:
                productive = session_hours
                idle = 0
        
        summary_dict[emp_id]['total_hours'] += session_hours
        summary_dict[emp_id]['total_productive_hours'] += productive
        summary_dict[emp_id]['total_idle_hours'] += idle
        summary_dict[emp_id]['days_worked'].add(session.date)
        summary_dict[emp_id]['sessions_count'] += 1
    
    # Convert to list and calculate average hours per day
    summary_list = []
    for emp_id, data in summary_dict.items():
        days_count = len(data['days_worked'])
        avg_hours = data['total_hours'] / days_count if days_count > 0 else 0
        
        # Cap average hours at 24 per day (physically impossible to work more)
        # This prevents display of erroneous data like 622.8 hours in 1 day
        if avg_hours > 24:
            # Log a warning but still show the data (might indicate data entry error)
            avg_hours = 24.0
        
        # Ensure productive + idle equals total (final validation at aggregate level)
        total_productive_idle = data['total_productive_hours'] + data['total_idle_hours']
        total_hours = data['total_hours']
        if abs(total_productive_idle - total_hours) > 0.01:
            # Normalize to ensure consistency
            if total_productive_idle > 0:
                scale_factor = total_hours / total_productive_idle
                data['total_productive_hours'] = data['total_productive_hours'] * scale_factor
                data['total_idle_hours'] = data['total_idle_hours'] * scale_factor
            else:
                # If no productive/idle data, assign all to productive
                data['total_productive_hours'] = total_hours
                data['total_idle_hours'] = 0
        
        # Use standard required hours for the period (same for everyone)
        # This is based on the calendar period, not individual days worked
        required_hours = total_required_hours
        variance = data['total_hours'] - required_hours
        variance_percent = (variance / required_hours * 100) if required_hours > 0 else 0
        
        summary_list.append({
            'employee_name': data['employee_name'],
            'employee_id': data['employee_id'],
            'department': data['department'],
            'total_hours': data['total_hours'],
            'days_worked': days_count,
            'avg_hours_per_day': avg_hours,
            'productive_hours': data['total_productive_hours'],
            'idle_hours': data['total_idle_hours'],
            'sessions_count': data['sessions_count'],
            'incomplete_sessions': data['incomplete_sessions'],
            'stale_sessions': data['stale_sessions'],  # Excluded old incomplete sessions
            'required_hours': required_hours,  # Standard hours for the period (same for all employees)
            'variance_hours': variance,  # Difference (actual - required)
            'variance_percent': variance_percent,  # Percentage difference
        })
    
    # Sort by total hours (descending)
    summary_list.sort(key=lambda x: x['total_hours'], reverse=True)
    return summary_list


def _generate_summary_csv_report(work_sessions, start_date, end_date, filename_date):
    """Generate CSV summary report"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="TimeTracking_Summary_{filename_date}.csv"'
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow([
        'Employee Name', 'Employee ID', 'Department', 'Days Worked', 'Required Hours', 
        'Actual Hours', 'Variance', 'Variance %', 'Avg Hours/Day',
        'Productive Hours', 'Idle Hours', 'Sessions Count'
    ])
    
    # Get summary data
    summary_data = _get_summary_data(work_sessions, start_date=start_date, end_date=end_date)
    
    # Data rows
    for emp_summary in summary_data:
        writer.writerow([
            emp_summary['employee_name'],
            emp_summary['employee_id'],
            emp_summary['department'],
            emp_summary['days_worked'],
            f"{emp_summary['required_hours']:.2f}",
            f"{emp_summary['total_hours']:.2f}",
            f"{emp_summary['variance_hours']:.2f}",
            f"{emp_summary['variance_percent']:.1f}%",
            f"{emp_summary['avg_hours_per_day']:.2f}",
            f"{emp_summary['productive_hours']:.2f}",
            f"{emp_summary['idle_hours']:.2f}",
            emp_summary['sessions_count'],
        ])
    
    # Summary rows
    total_hours = sum([emp['total_hours'] for emp in summary_data])
    total_days = sum([emp['days_worked'] for emp in summary_data])
    total_sessions = sum([emp['sessions_count'] for emp in summary_data])
    total_incomplete = sum([emp['incomplete_sessions'] for emp in summary_data])
    total_stale = sum([emp['stale_sessions'] for emp in summary_data])
    
    writer.writerow([])
    writer.writerow(['TOTAL SUMMARY'])
    writer.writerow(['Period', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"])
    writer.writerow(['Total Employees', len(summary_data)])
    writer.writerow(['Total Hours Across All Employees', f'{total_hours:.2f}'])
    writer.writerow(['Total Sessions', total_sessions])
    writer.writerow(['Incomplete Sessions (Active)', total_incomplete])
    writer.writerow(['Stale Sessions Excluded', total_stale])
    writer.writerow(['', ''])
    writer.writerow(['NOTE: Incomplete sessions older than 24 hours are excluded'])
    writer.writerow(['from reports as they are likely abandoned.'])
    writer.writerow(['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    return response


def _generate_summary_excel_report(work_sessions, start_date, end_date, filename_date):
    """Generate Excel summary report"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="TimeTracking_Summary_{filename_date}.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Summary'
    
    # Define styles
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    summary_font = Font(bold=True, size=10)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths - expanded for new columns
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 18
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 14
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 12
    worksheet.column_dimensions['H'].width = 12
    worksheet.column_dimensions['I'].width = 14
    worksheet.column_dimensions['J'].width = 14
    worksheet.column_dimensions['K'].width = 12
    worksheet.column_dimensions['L'].width = 14
    
    # Header row
    headers = [
        'Employee Name', 'Employee ID', 'Department', 'Days Worked', 'Required Hours', 
        'Actual Hours', 'Variance', 'Variance %', 'Avg Hours/Day',
        'Productive Hours', 'Idle Hours', 'Sessions Count'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Get summary data
    summary_data = _get_summary_data(work_sessions, start_date=start_date, end_date=end_date)
    
    # Data rows
    current_row = 2
    for emp_summary in summary_data:
        cells = [
            emp_summary['employee_name'],
            emp_summary['employee_id'],
            emp_summary['department'],
            emp_summary['days_worked'],
            f"{emp_summary['required_hours']:.2f}",
            f"{emp_summary['total_hours']:.2f}",
            f"{emp_summary['variance_hours']:.2f}",
            f"{emp_summary['variance_percent']:.1f}%",
            f"{emp_summary['avg_hours_per_day']:.2f}",
            f"{emp_summary['productive_hours']:.2f}",
            f"{emp_summary['idle_hours']:.2f}",
            emp_summary['sessions_count'],
        ]
        
        for col, value in enumerate(cells, 1):
            cell = worksheet.cell(row=current_row, column=col)
            cell.value = value
            cell.border = border
            if col in [4, 5, 6, 7, 8, 9]:  # Numeric columns - center alignment
                cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
    
    # Summary section
    summary_row = current_row + 2
    total_hours = sum([emp['total_hours'] for emp in summary_data])
    total_sessions = sum([emp['sessions_count'] for emp in summary_data])
    total_incomplete = sum([emp['incomplete_sessions'] for emp in summary_data])
    total_stale = sum([emp['stale_sessions'] for emp in summary_data])
    
    summary_data_rows = [
        ['TOTAL SUMMARY', ''],
        ['Period', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
        ['Total Employees', len(summary_data)],
        ['Total Hours Across All Employees', f'{total_hours:.2f}'],
        ['Total Sessions', total_sessions],
        ['Total In-Progress Sessions', total_incomplete],
        ['Stale Sessions Excluded', total_stale],
        ['', ''],
        ['NOTE', 'Incomplete sessions older than 24 hours are excluded'],
        ['', 'as they are likely abandoned/orphaned sessions'],
        ['Generated', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for idx, row_data in enumerate(summary_data_rows, 0):
        for col, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=summary_row + idx, column=col)
            cell.value = value
            if idx == 0:  # Title row
                cell.font = summary_font
                cell.fill = summary_fill
            cell.border = border
    
    workbook.save(response)
    return response


def _generate_summary_pdf_report(work_sessions, start_date, end_date, filename_date):
    """Generate PDF summary report"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="TimeTracking_Summary_{filename_date}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=12,
        alignment=1  # Center
    )
    story.append(Paragraph('Time Tracking Summary Report', title_style))
    
    # Period info
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12
    )
    period_text = f"Report Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    story.append(Paragraph(period_text, info_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data - with Required vs Actual Hours
    table_data = [[
        'Employee Name', 'Employee ID', 'Days', 'Required', 'Actual',
        'Variance', 'Var %', 'Avg/Day', 'Productive', 'Sessions'
    ]]
    
    # Get summary data
    summary_data = _get_summary_data(work_sessions, start_date=start_date, end_date=end_date)
    
    for emp_summary in summary_data:
        table_data.append([
            emp_summary['employee_name'],
            emp_summary['employee_id'],
            str(emp_summary['days_worked']),
            f"{emp_summary['required_hours']:.1f}",
            f"{emp_summary['total_hours']:.1f}",
            f"{emp_summary['variance_hours']:.1f}",
            f"{emp_summary['variance_percent']:.0f}%",
            f"{emp_summary['avg_hours_per_day']:.1f}",
            f"{emp_summary['productive_hours']:.1f}",
            str(emp_summary['sessions_count']),
        ])
    
    # Create table - reduced column widths for new layout
    table = Table(table_data, colWidths=[1.2*inch, 0.8*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.7*inch, 0.8*inch, 0.7*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Summary section
    total_hours = sum([emp['total_hours'] for emp in summary_data])
    total_sessions = sum([emp['sessions_count'] for emp in summary_data])
    total_incomplete = sum([emp['incomplete_sessions'] for emp in summary_data])
    total_stale = sum([emp['stale_sessions'] for emp in summary_data])
    
    summary_text = f"""
    <b>Summary Statistics:</b><br/>
    Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}<br/>
    Total Employees: {len(summary_data)}<br/>
    Total Hours Across All Employees: {total_hours:.2f}<br/>
    Total Sessions: {total_sessions}<br/>
    Total In-Progress Sessions: {total_incomplete}<br/>
    Stale Sessions Excluded: {total_stale}<br/>
    <br/>
    <b><i>⚠ Data Quality Note:</i></b><br/>
    Incomplete sessions older than 24 hours are excluded from reports as they are likely 
    abandoned/orphaned sessions with no punch_out time. These would artificially inflate 
    reported hours. In-progress sessions show real-time hours calculated from current time 
    and update dynamically as employees continue working.<br/>
    Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    return response


@login_required
@user_passes_test(is_admin_user)
def admin_reports(request):
    """Wrapper view for CRM admin reports - redirects to CRM reports view"""
    from crm.views import admin_reports as crm_admin_reports
    return crm_admin_reports(request)


@login_required
@user_passes_test(is_admin_user)
def admin_leave_redirect(request):
    return redirect('crm:admin_leave_list')
