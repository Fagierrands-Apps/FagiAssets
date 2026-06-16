from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .decorators import admin_required, sales_required, call_center_required, user_required, role_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from datetime import datetime, timedelta
from .models import (
    Customer, Employee, AssetCustomerAssignment, 
    Lead, CustomerNote, Notification, Department,
    TimeEntry, WorkSession, EmployeeKPI, Task, Communication, MissedCall,
    MoneyRequest, ActivityLog, IdlePeriod, ActivitySession, MonitoringSettings,
    HandlerReport
)
from assets.models import Asset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json


@role_required(['admin', 'sales', 'call_center'])
@login_required
def crm_dashboard(request):
    """Main CRM dashboard"""
    # Redirect non-admin users to employee portal
    try:
        employee_profile = request.user.employee_profile
        if employee_profile.role != 'admin':
            return redirect('crm:employee_dashboard')
    except Employee.DoesNotExist:
        pass  # Continue to main dashboard if no employee profile

    # Get summary statistics
    total_customers = Customer.objects.filter(status='active').count()
    total_leads = Lead.objects.exclude(status__in=['won', 'lost']).count()
    total_assignments = AssetCustomerAssignment.objects.filter(is_active=True).count()
    
    # Recent activity
    recent_customers = Customer.objects.all()[:5]
    recent_assignments = AssetCustomerAssignment.objects.select_related('asset', 'customer').all()[:5]
    recent_leads = Lead.objects.all()[:5]
    
    # Monthly stats
    current_month = timezone.now().replace(day=1)
    new_customers_this_month = Customer.objects.filter(created_at__gte=current_month).count()
    new_leads_this_month = Lead.objects.filter(created_at__gte=current_month).count()
    
    # Revenue from assignments (if applicable)
    monthly_revenue = AssetCustomerAssignment.objects.filter(
        is_active=True,
        monthly_fee__isnull=False
    ).aggregate(total=Sum('monthly_fee'))['total'] or 0
    
    context = {
        'stats': {
            'total_customers': total_customers,
            'total_leads': total_leads,
            'total_assignments': total_assignments,
            'new_customers_this_month': new_customers_this_month,
            'new_leads_this_month': new_leads_this_month,
            'monthly_revenue': monthly_revenue,
        },
        'recent_customers': recent_customers,
        'recent_assignments': recent_assignments,
        'recent_leads': recent_leads,
    }
    
    return render(request, 'crm/dashboard.html', context)


@role_required(['sales', 'call_center', 'admin'])
@login_required
def customer_list(request):
    """List all customers with search and filtering"""
    customers = Customer.objects.all().select_related('assigned_employee__user')
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        customers = customers.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(company_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        customers = customers.filter(status=status)
    
    # Filter by assigned employee
    assigned_employee = request.GET.get('assigned_employee')
    if assigned_employee:
        customers = customers.filter(assigned_employee_id=assigned_employee)
    
    # Pagination
    paginator = Paginator(customers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get employees for filter dropdown
    employees = Employee.objects.filter(employment_status='active').select_related('user')
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'assigned_employee': assigned_employee,
        'employees': employees,
        'status_choices': Customer.STATUS_CHOICES,
    }
    
    return render(request, 'crm/customer_list.html', context)


@user_required
@login_required
def customer_detail(request, customer_id):
    """Customer detail view with assets and notes"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # Get customer's asset assignments
    assignments = AssetCustomerAssignment.objects.filter(
        customer=customer
    ).select_related('asset', 'assigned_by').order_by('-assigned_date')
    
    # Get customer notes
    notes = customer.notes.all()[:10]
    
    # Get available assets for assignment
    available_assets = Asset.objects.filter(status='active').exclude(
        id__in=assignments.filter(is_active=True).values_list('asset_id', flat=True)
    )
    
    context = {
        'customer': customer,
        'assignments': assignments,
        'notes': notes,
        'available_assets': available_assets,
    }
    
    return render(request, 'crm/customer_detail.html', context)


@role_required(['sales', 'call_center', 'admin'])
@login_required
def customer_create(request):
    """Create new customer"""
    if request.method == 'POST':
        try:
            # Get employee profile if exists
            assigned_employee = None
            if request.POST.get('assigned_employee'):
                assigned_employee = Employee.objects.get(id=request.POST['assigned_employee'])
            
            customer = Customer.objects.create(
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name'],
                company_name=request.POST.get('company_name', ''),
                customer_type=request.POST.get('customer_type', 'individual'),
                email=request.POST['email'],
                phone=request.POST['phone'],
                alternate_phone=request.POST.get('alternate_phone', ''),
                address_line1=request.POST['address_line1'],
                address_line2=request.POST.get('address_line2', ''),
                city=request.POST['city'],
                state=request.POST['state'],
                postal_code=request.POST['postal_code'],
                country=request.POST.get('country', 'USA'),
                assigned_employee=assigned_employee,
                status='active'
            )
            
            messages.success(request, f'Customer {customer.full_name} created successfully!')
            return redirect('crm:customer_detail', customer_id=customer.id)
            
        except Exception as e:
            messages.error(request, f'Error creating customer: {str(e)}')
    
    # Get employees for assignment dropdown
    employees = Employee.objects.filter(employment_status='active').select_related('user')
    
    context = {
        'employees': employees,
        'customer_type_choices': Customer.CUSTOMER_TYPE_CHOICES,
    }
    
    return render(request, 'crm/customer_form.html', context)


@role_required(['sales', 'call_center', 'admin'])
@login_required
def assign_asset(request):
    """Assign asset to customer"""
    if request.method == 'POST':
        try:
            asset_id = request.POST.get('asset_id')
            customer_id = request.POST.get('customer_id')
            assignment_type = request.POST.get('assignment_type', 'owned')
            
            asset = get_object_or_404(Asset, id=asset_id)
            customer = get_object_or_404(Customer, id=customer_id)
            
            # Check for existing active assignment
            existing = AssetCustomerAssignment.objects.filter(
                asset=asset,
                customer=customer,
                is_active=True
            ).exists()
            
            if existing:
                messages.error(request, 'An active assignment already exists between this asset and customer.')
            else:
                assignment = AssetCustomerAssignment.objects.create(
                    asset=asset,
                    customer=customer,
                    assignment_type=assignment_type,
                    assigned_by=request.user,
                    contract_number=request.POST.get('contract_number', ''),
                    monthly_fee=request.POST.get('monthly_fee') or None,
                    total_value=request.POST.get('total_value') or None,
                    notes=request.POST.get('notes', ''),
                )
                
                # Create notification
                Notification.objects.create(
                    user=request.user,
                    notification_type='success',
                    title='Asset Assigned',
                    message=f'Asset {asset.asset_tag} assigned to {customer.full_name}',
                    customer=customer,
                    asset=asset
                )
                
                messages.success(request, f'Successfully assigned {asset.asset_tag} to {customer.full_name}')
                return redirect('crm:customer_detail', customer_id=customer.id)
        
        except Exception as e:
            messages.error(request, f'Error creating assignment: {str(e)}')
    
    return redirect('crm:customer_list')


@role_required(['sales', 'call_center', 'admin'])
@login_required
def lead_list(request):
    """List all leads"""
    leads = Lead.objects.all().select_related('assigned_employee__user')
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        leads = leads.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(company_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        leads = leads.filter(status=status)
    
    # Pagination
    paginator = Paginator(leads, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'status_choices': Lead.STATUS_CHOICES,
    }
    
    return render(request, 'crm/lead_list.html', context)


@role_required(['sales', 'call_center', 'admin'])
@login_required
def lead_create(request):
    """Create new lead"""
    if request.method == 'POST':
        try:
            lead = Lead.objects.create(
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                email=request.POST.get('email'),
                phone=request.POST.get('phone'),
                company=request.POST.get('company'),
                source=request.POST.get('source', 'website'),
                status=request.POST.get('status', 'new'),
                notes=request.POST.get('notes', ''),
                assigned_to_id=request.POST.get('assigned_to') if request.POST.get('assigned_to') else None
            )
            
            # Create notification
            Notification.objects.create(
                user=request.user,
                notification_type='success',
                title='New Lead Created',
                message=f'Lead {lead.full_name} has been created successfully.'
            )
            
            messages.success(request, f'Lead {lead.full_name} created successfully!')
            return redirect('crm:lead_detail', lead_id=lead.id)
            
        except Exception as e:
            messages.error(request, f'Error creating lead: {str(e)}')
    
    # Get employees for assignment
    employees = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'employees': employees,
        'source_choices': Lead.SOURCE_CHOICES,
        'status_choices': Lead.STATUS_CHOICES,
    }
    
    return render(request, 'crm/lead_create.html', context)


@role_required(['sales', 'call_center', 'admin'])
@login_required
def lead_detail(request, lead_id):
    """Lead detail view"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    context = {
        'lead': lead,
    }
    
    return render(request, 'crm/lead_detail.html', context)


@role_required(['sales', 'call_center', 'admin'])
@login_required
def convert_lead(request, lead_id):
    """Convert lead to customer"""
    if request.method == 'POST':
        lead = get_object_or_404(Lead, id=lead_id)
        
        try:
            customer = lead.convert_to_customer()
            
            # Create notification
            Notification.objects.create(
                user=request.user,
                notification_type='success',
                title='Lead Converted',
                message=f'Lead {lead.full_name} converted to customer',
                customer=customer,
                lead=lead
            )
            
            messages.success(request, f'Lead {lead.full_name} successfully converted to customer!')
            return redirect('crm:customer_detail', customer_id=customer.id)
            
        except Exception as e:
            messages.error(request, f'Error converting lead: {str(e)}')
    
    return redirect('crm:lead_detail', lead_id=lead_id)


@sales_required
@login_required
def assignment_list(request):
    """List asset assignments"""
    assignments = AssetCustomerAssignment.objects.all().select_related(
        'asset', 'customer', 'assigned_by'
    )
    
    # Filter by assignment type
    assignment_type = request.GET.get('assignment_type')
    if assignment_type:
        assignments = assignments.filter(assignment_type=assignment_type)
    
    # Filter by active status
    is_active = request.GET.get('is_active')
    if is_active is not None:
        assignments = assignments.filter(is_active=is_active.lower() == 'true')
    
    # Pagination
    paginator = Paginator(assignments, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'assignment_type': assignment_type,
        'is_active': is_active,
        'assignment_type_choices': AssetCustomerAssignment.ASSIGNMENT_TYPE_CHOICES,
    }
    
    return render(request, 'crm/assignment_list.html', context)


@login_required
def notifications_view(request):
    """View user notifications"""
    notifications = request.user.notifications.filter(is_dismissed=False)
    
    # Mark as read if requested
    if request.GET.get('mark_read'):
        notifications.update(is_read=True)
    
    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'crm/notifications.html', context)


@login_required
def dismiss_notification(request, notification_id):
    """Dismiss a notification"""
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_dismissed = True
    notification.save()
    
    return JsonResponse({'status': 'success'})


# API Views for AJAX requests
@login_required
def customer_search_api(request):
    """API endpoint for customer search"""
    search = request.GET.get('q', '')
    customers = Customer.objects.filter(
        Q(first_name__icontains=search) |
        Q(last_name__icontains=search) |
        Q(company_name__icontains=search)
    )[:10]
    
    results = []
    for customer in customers:
        results.append({
            'id': customer.id,
            'text': str(customer),
            'email': customer.email,
            'phone': customer.phone,
        })
    
    return JsonResponse({'results': results})


@login_required
def asset_search_api(request):
    """API endpoint for asset search"""
    search = request.GET.get('q', '')
    assets = Asset.objects.filter(
        Q(asset_tag__icontains=search) |
        Q(name__icontains=search),
        status='active'
    )[:10]
    
    results = []
    for asset in assets:
        results.append({
            'id': asset.id,
            'text': f"{asset.asset_tag} - {asset.name}",
            'model': str(asset.model),
            'status': asset.status,
        })
    
    return JsonResponse({'results': results})


# Employee Time Tracking Views
@login_required
def employee_dashboard(request):
    """Employee-specific dashboard with time tracking and tasks"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found. Please contact administrator.')
        return redirect('crm:dashboard')
    
    # Get current status
    current_status = employee.get_current_status()
    today_hours = employee.get_today_hours()
    pending_tasks = employee.get_pending_tasks_count()
    
    # Get today's time entries
    today = timezone.now().date()
    today_entries = employee.time_entries.filter(timestamp__date=today)
    
    # Get recent tasks
    recent_tasks = employee.assigned_tasks.filter(
        status__in=['pending', 'in_progress']
    ).order_by('-created_at')[:5]
    
    # Get recent communications
    recent_communications = employee.communications.all()[:5]
    
    # Get monthly KPIs
    monthly_kpis = employee.get_monthly_kpi_summary()
    
    context = {
        'employee': employee,
        'current_status': current_status,
        'today_hours': today_hours,
        'pending_tasks': pending_tasks,
        'today_entries': today_entries,
        'recent_tasks': recent_tasks,
        'recent_communications': recent_communications,
        'monthly_kpis': monthly_kpis,
    }
    
    return render(request, 'crm/employee_dashboard.html', context)


@login_required
def punch_in_out(request):
    """Handle punch in/out functionality"""
    import logging
    logger = logging.getLogger(__name__)
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found. Please contact administrator.')
        return redirect('crm:employee_dashboard')

    if request.method == 'POST':
        try:
            entry_type = request.POST.get('entry_type')
            location = request.POST.get('location', '')
            notes = request.POST.get('notes', '')

            # Get client IP
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            # Create time entry
            time_entry = TimeEntry.objects.create(
                employee=employee,
                entry_type=entry_type,
                location=location,
                notes=notes,
                ip_address=ip_address,
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )

            # Handle work session logic
            today = timezone.now().date()

            if entry_type == 'punch_in':
                # Create or update work session
                work_session, created = WorkSession.objects.get_or_create(
                    employee=employee,
                    date=today,
                    defaults={'punch_in': time_entry.timestamp}
                )
                if not created and not work_session.punch_in:
                    work_session.punch_in = time_entry.timestamp
                    work_session.save()

                messages.success(request, f'Successfully punched in at {time_entry.timestamp.strftime("%H:%M")}')

            elif entry_type == 'punch_out':
                try:
                    work_session = WorkSession.objects.get(employee=employee, date=today)
                    work_session.punch_out = time_entry.timestamp
                    work_session.calculate_hours()
                    messages.success(request, f'Successfully punched out at {time_entry.timestamp.strftime("%H:%M")}. Total hours: {work_session.worked_hours}')
                except WorkSession.DoesNotExist:
                    messages.warning(request, 'No punch-in record found for today. Please contact administrator.')

            elif entry_type in ['break_start', 'break_end']:
                messages.success(request, f'Break {entry_type.split("_")[1]} recorded at {time_entry.timestamp.strftime("%H:%M")}')

            # Redirect back to the page after POST
            return redirect('crm:punch_in_out')

        except Exception as e:
            logger.error(f'Error recording time entry: {str(e)}', exc_info=True)
            messages.error(request, f'Error recording time entry: {str(e)}')
            return redirect('crm:punch_in_out')

    # GET request - render the template
    current_status = employee.get_current_status()
    today_hours = employee.get_today_hours()

    # Calculate break hours from time entries
    break_entries = employee.time_entries.filter(
        entry_type__in=['break_start', 'break_end'],
        timestamp__date=timezone.now().date()
    ).order_by('timestamp')

    break_hours = 0
    last_break_start = None
    for entry in break_entries:
        if entry.entry_type == 'break_start':
            last_break_start = entry.timestamp
        elif entry.entry_type == 'break_end' and last_break_start:
            break_duration = (entry.timestamp - last_break_start).total_seconds() / 3600
            break_hours += break_duration
            last_break_start = None

    total_entries = employee.time_entries.filter(timestamp__date=timezone.now().date()).count()
    today_entries = employee.time_entries.filter(timestamp__date=timezone.now().date())[:5]

    # Format current status for display
    status_display_map = {
        'not_punched_in': 'Not Punched In',
        'punched_in': 'Working',
        'on_break': 'On Break',
        'punched_out': 'Punched Out',
    }
    current_status_display = status_display_map.get(current_status, current_status.title().replace('_', ' '))

    # Check if it's lunch time (1 PM - 2 PM, Monday-Saturday)
    now = timezone.now()
    is_lunch_time = (
        now.hour == 13 and  # 1 PM hour (13:00-13:59)
        now.weekday() < 6  # Monday (0) to Saturday (5), exclude Sunday (6)
    )

    context = {
        'employee': employee,
        'current_status': current_status_display,
        'today_hours': today_hours,
        'break_hours': break_hours,
        'total_entries': total_entries,
        'today_entries': today_entries,
        'is_lunch_time': is_lunch_time,
    }

    return render(request, 'crm/punch_in_out.html', context)


@login_required
def employee_timesheet(request):
    """View employee timesheet"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    # Get date range from request or default to current month
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get work sessions for the period
    work_sessions = employee.work_sessions.filter(
        date__range=[start_date, end_date]
    ).order_by('-date')
    
    # Get time entries for the period
    time_entries = employee.time_entries.filter(
        timestamp__date__range=[start_date, end_date]
    ).order_by('-timestamp')
    
    # Calculate totals
    total_hours = sum(session.worked_hours for session in work_sessions)
    total_days = work_sessions.filter(is_complete=True).count()
    avg_hours_per_day = total_hours / total_days if total_days > 0 else 0
    
    context = {
        'employee': employee,
        'work_sessions': work_sessions,
        'time_entries': time_entries,
        'start_date': start_date,
        'end_date': end_date,
        'total_hours': total_hours,
        'total_days': total_days,
        'avg_hours_per_day': avg_hours_per_day,
    }
    
    return render(request, 'crm/employee_timesheet.html', context)


@login_required
def employee_kpis(request):
    """View employee KPIs and performance metrics"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    # Get date range from request or default to current month
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    if not month or not year:
        today = timezone.now()
        month = today.month
        year = today.year
    else:
        month = int(month)
        year = int(year)
    
    # Get KPIs for the period
    kpis = employee.kpis.filter(
        period_start__month=month,
        period_start__year=year
    ).order_by('kpi_type')
    
    # Get KPI summary
    kpi_summary = employee.get_monthly_kpi_summary(month, year)
    
    # Get available KPI types for adding new ones
    kpi_types = EmployeeKPI.KPI_TYPE_CHOICES
    
    context = {
        'employee': employee,
        'kpis': kpis,
        'kpi_summary': kpi_summary,
        'kpi_types': kpi_types,
        'selected_month': month,
        'selected_year': year,
    }
    
    return render(request, 'crm/employee_kpis.html', context)


@login_required
def employee_tasks(request):
    """View and manage employee tasks"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    # Get tasks with filtering
    tasks = employee.assigned_tasks.all()
    
    status_filter = request.GET.get('status')
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    priority_filter = request.GET.get('priority')
    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)
    
    # Pagination
    paginator = Paginator(tasks, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'employee': employee,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
    }
    
    return render(request, 'crm/employee_tasks.html', context)


@login_required
def update_task_status(request, task_id):
    """Update task status"""
    if request.method == 'POST':
        try:
            employee = request.user.employee_profile
            task = get_object_or_404(Task, id=task_id, assigned_to=employee)
            
            new_status = request.POST.get('status')
            actual_hours = request.POST.get('actual_hours')
            
            task.status = new_status
            if actual_hours:
                task.actual_hours = actual_hours
            
            if new_status == 'completed':
                task.mark_completed()
            
            task.save()
            
            messages.success(request, f'Task "{task.title}" status updated to {task.get_status_display()}')
            return JsonResponse({'status': 'success'})
            
        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee profile not found'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


@login_required
def employee_communications(request):
    """View and manage employee communications"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    # Get communications with filtering
    communications = employee.communications.all()
    
    comm_type_filter = request.GET.get('type')
    if comm_type_filter:
        communications = communications.filter(communication_type=comm_type_filter)
    
    # Pagination
    paginator = Paginator(communications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'employee': employee,
        'page_obj': page_obj,
        'comm_type_filter': comm_type_filter,
        'communication_types': Communication.COMMUNICATION_TYPE_CHOICES,
    }
    
    return render(request, 'crm/employee_communications.html', context)


@login_required
def money_request(request):
    """Submit money requests to the company"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')

    if request.method == 'POST':
        from_list = request.POST.getlist('money_request_from[]')
        to_list = request.POST.getlist('money_request_to[]')
        date_list = request.POST.getlist('money_request_date[]')
        note = request.POST.get('money_request_notes', '').strip()

        valid_rows = []
        for from_entity, to_entity, date_str in zip(from_list, to_list, date_list):
            from_entity = from_entity.strip()
            to_entity = to_entity.strip()
            if not from_entity or not to_entity or not date_str:
                continue

            try:
                request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                continue

            valid_rows.append((from_entity, to_entity, request_date))

        if not valid_rows:
            messages.error(request, 'Please add at least one complete request line with a valid date.')
        else:
            for from_entity, to_entity, request_date in valid_rows:
                MoneyRequest.objects.create(
                    employee=employee,
                    from_entity=from_entity,
                    to_entity=to_entity,
                    request_date=request_date,
                    notes=note,
                )

            request_count = len(valid_rows)
            total_amount = request_count * 100
            messages.success(
                request,
                f'Funds request submitted successfully! {request_count} line{"s" if request_count != 1 else ""} totalling KSh {total_amount}.'
            )
            return redirect('crm:money_request')

    history = MoneyRequest.objects.filter(employee=employee).order_by('-created_at')

    context = {
        'employee': employee,
        'money_requests': history,
    }
    return render(request, 'crm/money_request.html', context)


@login_required
def money_request_list(request):
    """Allow accountants to review submitted fund requests"""
    try:
        employee = request.user.employee_profile
    except (Employee.DoesNotExist, AttributeError):
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')

    money_requests = MoneyRequest.objects.select_related('employee__user', 'processed_by__user')
    status_filter = request.GET.get('status')
    if status_filter:
        money_requests = money_requests.filter(status=status_filter)

    context = {
        'employee': employee,
        'money_requests': money_requests,
        'status_filter': status_filter,
        'status_choices': MoneyRequest.STATUS_CHOICES,
    }
    return render(request, 'crm/money_request_list.html', context)


@role_required(['accountant', 'admin'])
@login_required
def money_request_decision(request, request_id):
    """Approve or reject a money request"""
    if request.method != 'POST':
        return redirect('crm:money_request_list')

    try:
        employee = request.user.employee_profile
    except (Employee.DoesNotExist, AttributeError):
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')

    money_request = get_object_or_404(MoneyRequest, id=request_id)

    action = request.POST.get('action', '').strip() if request.POST.get('action') else None
    decision_notes = request.POST.get('decision_notes', '').strip()

    if action not in ['approve', 'reject']:
        messages.error(request, 'Invalid action. Please use the approve or reject buttons.')
        return redirect('crm:money_request_list')

    new_status = 'approved' if action == 'approve' else 'rejected'
    if money_request.status == new_status:
        messages.info(request, f'Request is already marked as {money_request.get_status_display().lower()}.')
        return redirect('crm:money_request_list')

    money_request.status = new_status
    money_request.decision_notes = decision_notes
    money_request.processed_by = employee
    money_request.processed_at = timezone.now()
    money_request.save()

    messages.success(
        request,
        f'Funds request { "approved" if new_status == "approved" else "rejected" } successfully.'
    )
    return redirect('crm:money_request_list')


@login_required
def add_communication(request):
    """Add new communication record"""
    if request.method == 'POST':
        try:
            employee = request.user.employee_profile
            
            communication = Communication.objects.create(
                employee=employee,
                communication_type=request.POST['communication_type'],
                direction=request.POST.get('direction', 'outbound'),
                subject=request.POST['subject'],
                content=request.POST['content'],
                contact_name=request.POST.get('contact_name', ''),
                contact_email=request.POST.get('contact_email', ''),
                contact_phone=request.POST.get('contact_phone', ''),
                duration_minutes=request.POST.get('duration_minutes') or None,
                requires_followup=request.POST.get('requires_followup') == 'on',
                followup_date=request.POST.get('followup_date') or None,
                is_sourced=request.POST.get('is_sourced') == 'on',
            )
            
            # Link to customer or lead if provided
            customer_id = request.POST.get('customer_id')
            if customer_id:
                communication.customer_id = customer_id
                communication.save()
            
            lead_id = request.POST.get('lead_id')
            if lead_id:
                communication.lead_id = lead_id
                communication.save()
            
            messages.success(request, 'Communication record added successfully!')
            return redirect('crm:employee_communications')
            
        except Employee.DoesNotExist:
            messages.error(request, 'Employee profile not found.')
        except Exception as e:
            messages.error(request, f'Error adding communication: {str(e)}')
    
    # Get customers and leads for dropdowns
    customers = Customer.objects.filter(status='active')[:100]
    leads = Lead.objects.exclude(status__in=['won', 'lost'])[:100]
    
    context = {
        'communication_types': Communication.COMMUNICATION_TYPE_CHOICES,
        'direction_choices': Communication.DIRECTION_CHOICES,
        'customers': customers,
        'leads': leads,
    }
    
    return render(request, 'crm/add_communication.html', context)


# Manager Task Management Views
@login_required
def manager_task_list(request):
    """View all tasks for managers"""
    try:
        employee = request.user.employee_profile
        if not employee.is_manager:
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('crm:employee_dashboard')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    # Get all tasks (managers can see all tasks)
    tasks = Task.objects.select_related('assigned_to__user', 'assigned_by').all()
    
    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    priority_filter = request.GET.get('priority')
    if priority_filter:
        tasks = tasks.filter(priority=priority_filter)
    
    employee_filter = request.GET.get('employee')
    if employee_filter:
        tasks = tasks.filter(assigned_to_id=employee_filter)
    
    # Order by priority and due date
    tasks = tasks.order_by('-priority', 'due_date', '-created_at')
    
    # Pagination
    paginator = Paginator(tasks, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get employees for filter
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    
    # Get task statistics
    total_tasks = Task.objects.count()
    pending_tasks = Task.objects.filter(status='pending').count()
    in_progress_tasks = Task.objects.filter(status='in_progress').count()
    completed_tasks = Task.objects.filter(status='completed').count()
    
    context = {
        'employee': employee,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'employee_filter': employee_filter,
        'employees': employees,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'total_tasks': total_tasks,
        'pending_tasks': pending_tasks,
        'in_progress_tasks': in_progress_tasks,
        'completed_tasks': completed_tasks,
    }
    
    return render(request, 'crm/manager_task_list.html', context)


@login_required
def manager_create_task(request):
    """Create a new task (managers only)"""
    try:
        employee = request.user.employee_profile
        if not employee.is_manager:
            messages.error(request, 'You do not have permission to create tasks.')
            return redirect('crm:employee_dashboard')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    if request.method == 'POST':
        try:
            # Create task
            task = Task.objects.create(
                title=request.POST['title'],
                description=request.POST['description'],
                assigned_to_id=request.POST['assigned_to'],
                assigned_by=request.user,
                priority=request.POST.get('priority', 'medium'),
                status='pending',
                due_date=request.POST.get('due_date') or None,
                estimated_hours=request.POST.get('estimated_hours') or None,
            )
            
            # Link to customer, lead, or asset if provided
            customer_id = request.POST.get('customer')
            if customer_id:
                task.customer_id = customer_id
            
            lead_id = request.POST.get('lead')
            if lead_id:
                task.lead_id = lead_id
            
            asset_id = request.POST.get('asset')
            if asset_id:
                task.asset_id = asset_id
            
            task.save()
            
            messages.success(request, f'Task "{task.title}" created and assigned to {task.assigned_to.full_name}')
            return redirect('crm:manager_task_list')
            
        except Exception as e:
            messages.error(request, f'Error creating task: {str(e)}')
    
    # Get data for dropdowns
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    customers = Customer.objects.filter(status='active').order_by('first_name', 'last_name')[:100]
    leads = Lead.objects.exclude(status__in=['won', 'lost']).order_by('company_name')[:100]
    assets = Asset.objects.filter(status='active').order_by('name')[:100]
    
    context = {
        'employee': employee,
        'employees': employees,
        'customers': customers,
        'leads': leads,
        'assets': assets,
        'priority_choices': Task.PRIORITY_CHOICES,
    }
    
    return render(request, 'crm/manager_create_task.html', context)


@login_required
def manager_edit_task(request, task_id):
    """Edit an existing task (managers only)"""
    try:
        employee = request.user.employee_profile
        if not employee.is_manager:
            messages.error(request, 'You do not have permission to edit tasks.')
            return redirect('crm:employee_dashboard')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('crm:dashboard')
    
    task = get_object_or_404(Task, id=task_id)
    
    if request.method == 'POST':
        try:
            # Update task
            task.title = request.POST['title']
            task.description = request.POST['description']
            task.assigned_to_id = request.POST['assigned_to']
            task.priority = request.POST.get('priority', 'medium')
            task.status = request.POST.get('status', task.status)
            task.due_date = request.POST.get('due_date') or None
            task.estimated_hours = request.POST.get('estimated_hours') or None
            task.actual_hours = request.POST.get('actual_hours') or None
            
            # Update related objects
            customer_id = request.POST.get('customer')
            task.customer_id = customer_id if customer_id else None
            
            lead_id = request.POST.get('lead')
            task.lead_id = lead_id if lead_id else None
            
            asset_id = request.POST.get('asset')
            task.asset_id = asset_id if asset_id else None
            
            # Mark as completed if status changed to completed
            if task.status == 'completed' and not task.completed_at:
                task.mark_completed()
            
            task.save()
            
            messages.success(request, f'Task "{task.title}" updated successfully')
            return redirect('crm:manager_task_list')
            
        except Exception as e:
            messages.error(request, f'Error updating task: {str(e)}')
    
    # Get data for dropdowns
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    customers = Customer.objects.filter(status='active').order_by('first_name', 'last_name')[:100]
    leads = Lead.objects.exclude(status__in=['won', 'lost']).order_by('company_name')[:100]
    assets = Asset.objects.filter(status='active').order_by('name')[:100]
    
    context = {
        'employee': employee,
        'task': task,
        'employees': employees,
        'customers': customers,
        'leads': leads,
        'assets': assets,
        'priority_choices': Task.PRIORITY_CHOICES,
        'status_choices': Task.STATUS_CHOICES,
    }
    
    return render(request, 'crm/manager_edit_task.html', context)


@login_required
def manager_delete_task(request, task_id):
    """Delete a task (managers only)"""
    try:
        employee = request.user.employee_profile
        if not employee.is_manager:
            return JsonResponse({'status': 'error', 'message': 'Permission denied'})
    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Employee profile not found'})
    
    if request.method == 'POST':
        try:
            task = get_object_or_404(Task, id=task_id)
            task_title = task.title
            task.delete()
            messages.success(request, f'Task "{task_title}" deleted successfully')
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


# ============================================================================
# MISSED CALL VIEWS
# ============================================================================

@login_required
def log_missed_call(request):
    """Log a new missed call"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found')
        return redirect('crm:employee_dashboard')
    
    if request.method == 'POST':
        caller_name = request.POST.get('caller_name')
        caller_phone = request.POST.get('caller_phone')
        caller_company = request.POST.get('caller_company', '')
        reason_for_call = request.POST.get('reason_for_call', '')
        priority = request.POST.get('priority', 'medium')
        customer_id = request.POST.get('customer')
        lead_id = request.POST.get('lead')
        callback_scheduled = request.POST.get('callback_scheduled_at')
        
        # Create missed call
        missed_call = MissedCall.objects.create(
            employee=employee,
            caller_name=caller_name,
            caller_phone=caller_phone,
            caller_company=caller_company,
            reason_for_call=reason_for_call,
            priority=priority,
            customer_id=customer_id if customer_id else None,
            lead_id=lead_id if lead_id else None,
            callback_scheduled_at=callback_scheduled if callback_scheduled else None
        )
        
        messages.success(request, f'Missed call from {caller_name} logged successfully')
        return redirect('crm:missed_calls_list')
    
    # GET request - show form
    customers = Customer.objects.filter(status='active').order_by('first_name', 'last_name')
    leads = Lead.objects.exclude(status__in=['won', 'lost']).order_by('first_name', 'last_name')
    
    context = {
        'employee': employee,
        'customers': customers,
        'leads': leads,
        'priority_choices': MissedCall.PRIORITY_CHOICES,
    }
    
    return render(request, 'crm/log_missed_call.html', context)


@login_required
def missed_calls_list(request):
    """View all missed calls for the logged-in employee"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found')
        return redirect('crm:employee_dashboard')
    
    # Filter missed calls for this employee
    missed_calls = MissedCall.objects.filter(employee=employee)
    
    # Apply filters
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    
    if status_filter:
        missed_calls = missed_calls.filter(status=status_filter)
    if priority_filter:
        missed_calls = missed_calls.filter(priority=priority_filter)
    
    # Get statistics
    total_calls = missed_calls.count()
    pending_calls = missed_calls.filter(status='pending').count()
    completed_calls = missed_calls.filter(status='completed').count()
    overdue_calls = sum(1 for call in missed_calls.filter(status='pending') if call.is_overdue)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(missed_calls, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'employee': employee,
        'missed_calls': page_obj,
        'total_calls': total_calls,
        'pending_calls': pending_calls,
        'completed_calls': completed_calls,
        'overdue_calls': overdue_calls,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'status_choices': MissedCall.STATUS_CHOICES,
        'priority_choices': MissedCall.PRIORITY_CHOICES,
    }
    
    return render(request, 'crm/missed_calls_list.html', context)


@login_required
def edit_missed_call(request, call_id):
    """Edit a missed call"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found')
        return redirect('crm:employee_dashboard')
    
    missed_call = get_object_or_404(MissedCall, id=call_id, employee=employee)
    
    if request.method == 'POST':
        missed_call.caller_name = request.POST.get('caller_name')
        missed_call.caller_phone = request.POST.get('caller_phone')
        missed_call.caller_company = request.POST.get('caller_company', '')
        missed_call.reason_for_call = request.POST.get('reason_for_call', '')
        missed_call.priority = request.POST.get('priority', 'medium')
        missed_call.status = request.POST.get('status', 'pending')
        missed_call.follow_up_notes = request.POST.get('follow_up_notes', '')
        
        customer_id = request.POST.get('customer')
        lead_id = request.POST.get('lead')
        callback_scheduled = request.POST.get('callback_scheduled_at')
        
        missed_call.customer_id = customer_id if customer_id else None
        missed_call.lead_id = lead_id if lead_id else None
        missed_call.callback_scheduled_at = callback_scheduled if callback_scheduled else None
        
        # If marking as completed, set completion timestamp
        if missed_call.status == 'completed' and not missed_call.callback_completed_at:
            missed_call.callback_completed_at = timezone.now()
            missed_call.callback_completed_by = request.user
        
        missed_call.save()
        
        messages.success(request, 'Missed call updated successfully')
        return redirect('crm:missed_calls_list')
    
    # GET request - show form
    customers = Customer.objects.filter(status='active').order_by('first_name', 'last_name')
    leads = Lead.objects.exclude(status__in=['won', 'lost']).order_by('first_name', 'last_name')
    
    context = {
        'employee': employee,
        'missed_call': missed_call,
        'customers': customers,
        'leads': leads,
        'priority_choices': MissedCall.PRIORITY_CHOICES,
        'status_choices': MissedCall.STATUS_CHOICES,
    }
    
    return render(request, 'crm/edit_missed_call.html', context)


@login_required
def delete_missed_call(request, call_id):
    """Delete a missed call"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Employee profile not found'})
    
    if request.method == 'POST':
        try:
            missed_call = get_object_or_404(MissedCall, id=call_id, employee=employee)
            caller_name = missed_call.caller_name
            missed_call.delete()
            messages.success(request, f'Missed call from {caller_name} deleted successfully')
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


@admin_required
@login_required
def productivity_dashboard(request):
    """
    Productivity Dashboard for Managers
    Shows team activity monitoring and productivity metrics
    """
    # Date filter
    date_filter = request.GET.get('date', 'today')
    
    if date_filter == 'today':
        start_date = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = timezone.now()
    elif date_filter == 'yesterday':
        yesterday = timezone.now() - timedelta(days=1)
        start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_filter == 'week':
        start_date = timezone.now() - timedelta(days=7)
        end_date = timezone.now()
    elif date_filter == 'month':
        start_date = timezone.now() - timedelta(days=30)
        end_date = timezone.now()
    else:
        start_date = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = timezone.now()
    
    # Get all employees with monitoring enabled
    employees_data = []
    employees = Employee.objects.filter(
        monitoring_settings__enable_monitoring=True
    ).select_related('monitoring_settings')
    
    for employee in employees:
        # Get work sessions in date range
        work_sessions = WorkSession.objects.filter(
            employee=employee,
            punch_in__gte=start_date,
            punch_in__lte=end_date
        )
        
        # Calculate totals
        total_worked_hours = sum([ws.calculate_hours() or 0 for ws in work_sessions])
        total_productive_hours = sum([ws.productive_hours or 0 for ws in work_sessions])
        total_idle_hours = sum([ws.idle_hours or 0 for ws in work_sessions])
        
        # Get activity sessions
        activity_sessions = ActivitySession.objects.filter(
            employee=employee,
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        
        # Calculate average productivity score
        if activity_sessions.exists():
            avg_productivity = activity_sessions.aggregate(
                avg_score=Sum('productivity_score')
            )['avg_score'] / activity_sessions.count()
        else:
            avg_productivity = 0
        
        # Get idle periods
        idle_periods = IdlePeriod.objects.filter(
            employee=employee,
            start_time__gte=start_date,
            start_time__lte=end_date
        )
        
        # Get activity counts
        activities = ActivityLog.objects.filter(
            employee=employee,
            timestamp__gte=start_date,
            timestamp__lte=end_date
        )
        
        mouse_events = activities.filter(activity_type='mouse_move').count()
        click_events = activities.filter(activity_type='click').count()
        keyboard_events = activities.filter(activity_type='keyboard').count()
        
        # Current status
        current_session = WorkSession.objects.filter(
            employee=employee,
            punch_out__isnull=True
        ).first()
        
        if current_session:
            # Check if currently idle
            last_activity = ActivityLog.objects.filter(
                employee=employee
            ).order_by('-timestamp').first()
            
            if last_activity:
                idle_seconds = (timezone.now() - last_activity.timestamp).total_seconds()
                is_idle = idle_seconds > employee.monitoring_settings.idle_threshold
                status = 'idle' if is_idle else 'active'
            else:
                status = 'no_activity'
        else:
            status = 'offline'
        
        employees_data.append({
            'employee': employee,
            'total_worked_hours': round(total_worked_hours, 2),
            'total_productive_hours': round(total_productive_hours, 2),
            'total_idle_hours': round(total_idle_hours, 2),
            'avg_productivity': round(avg_productivity, 1),
            'idle_periods_count': idle_periods.count(),
            'mouse_events': mouse_events,
            'click_events': click_events,
            'keyboard_events': keyboard_events,
            'status': status,
            'current_session': current_session,
        })
    
    # Sort by productivity score
    employees_data.sort(key=lambda x: x['avg_productivity'], reverse=True)
    
    # Overall statistics
    total_employees = len(employees_data)
    active_employees = sum(1 for e in employees_data if e['status'] == 'active')
    idle_employees = sum(1 for e in employees_data if e['status'] == 'idle')
    offline_employees = sum(1 for e in employees_data if e['status'] == 'offline')
    
    total_productive_time = sum(e['total_productive_hours'] for e in employees_data)
    total_idle_time = sum(e['total_idle_hours'] for e in employees_data)
    
    if total_productive_time + total_idle_time > 0:
        overall_productivity = (total_productive_time / (total_productive_time + total_idle_time)) * 100
    else:
        overall_productivity = 0
    
    context = {
        'employees_data': employees_data,
        'date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'idle_employees': idle_employees,
        'offline_employees': offline_employees,
        'total_productive_time': round(total_productive_time, 2),
        'total_idle_time': round(total_idle_time, 2),
        'overall_productivity': round(overall_productivity, 1),
    }
    
    return render(request, 'crm/productivity_dashboard.html', context)


@login_required
def handler_reports(request):
    """Weekly reports list"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found')
        return redirect('crm:employee_dashboard')
    
    reports = HandlerReport.objects.filter(handler=employee).order_by('-week_start')
    
    context = {
        'reports': reports,
        'employee': employee,
    }
    
    return render(request, 'crm/handler_reports.html', context)


@login_required
def handler_report_detail(request, report_id):
    """View/edit weekly report"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found')
        return redirect('crm:employee_dashboard')
    
    report = get_object_or_404(HandlerReport, id=report_id, handler=employee)
    
    if request.method == 'POST':
        try:
            report.calls_made = int(request.POST.get('calls_made', 0))
            report.successful_calls = int(request.POST.get('successful_calls', 0))
            report.errands_made = int(request.POST.get('errands_made', 0))
            report.notes = request.POST.get('notes', '')
            report.save()
            
            messages.success(request, 'Report updated successfully!')
            return redirect('crm:handler_reports')
        except Exception as e:
            messages.error(request, f'Error updating report: {str(e)}')
    
    context = {
        'report': report,
        'employee': employee,
    }
    
    return render(request, 'crm/handler_report_detail.html', context)


@login_required
def handler_report_create(request):
    """Create new weekly report"""
    try:
        employee = request.user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found')
        return redirect('crm:employee_dashboard')
    
    if request.method == 'POST':
        try:
            from datetime import datetime, timedelta
            report_date = datetime.strptime(request.POST.get('report_date'), '%Y-%m-%d').date()
            
            existing = HandlerReport.objects.filter(handler=employee, week_start=report_date).exists()
            if existing:
                messages.error(request, 'A report already exists for this date.')
                return redirect('crm:handler_report_create')
            
            report = HandlerReport.objects.create(
                handler=employee,
                week_start=report_date,
                week_end=report_date,
                calls_made=int(request.POST.get('calls_made', 0)),
                successful_calls=int(request.POST.get('successful_calls', 0)),
                errands_made=int(request.POST.get('errands_made', 0)),
                weekly_target=employee.weekly_target,
                notes=request.POST.get('notes', '')
            )
            
            messages.success(request, 'Report created successfully!')
            return redirect('crm:handler_reports')
        except Exception as e:
            messages.error(request, f'Error creating report: {str(e)}')
    
    context = {
        'employee': employee,
    }
    
    return render(request, 'crm/handler_report_create.html', context)


@login_required
def admin_reports(request):
    """Admin and Accountant view - all daily reports from all employees"""
    # Check if user is staff or superuser (admin dashboard access)
    if not (request.user.is_staff or request.user.is_superuser):
        # Also check if user has employee profile with admin or accountant role
        try:
            employee = request.user.employee_profile
            if employee.role not in ['admin', 'accountant']:
                messages.error(request, 'Access denied. Only admins and accountants can view this page.')
                return redirect('crm:employee_dashboard')
        except (Employee.DoesNotExist, AttributeError):
            messages.error(request, 'Access denied. Only admins and accountants can view this page.')
            return redirect('crm:employee_dashboard')
    
    reports = HandlerReport.objects.all().select_related('handler__user').order_by('-week_start')
    
    search = request.GET.get('search')
    if search:
        reports = reports.filter(
            Q(handler__user__first_name__icontains=search) |
            Q(handler__user__last_name__icontains=search) |
            Q(handler__user__username__icontains=search)
        )
    
    date_filter = request.GET.get('date')
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            reports = reports.filter(week_start=filter_date)
        except ValueError:
            pass
    
    employee_filter = request.GET.get('employee')
    if employee_filter:
        reports = reports.filter(handler_id=employee_filter)
    
    paginator = Paginator(reports, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    employees = Employee.objects.filter(employment_status='active').order_by('user__first_name')
    
    # Check if this is a download request
    format_type = request.GET.get('format')
    if format_type == 'xlsx':
        return _generate_reports_excel(reports, search, date_filter, employee_filter)
    
    context = {
        'page_obj': page_obj,
        'reports': page_obj.object_list,
        'total_reports': paginator.count,
        'employees': employees,
        'search': search,
        'date_filter': date_filter,
        'employee_filter': employee_filter,
    }
    
    # Use different template for accountants (extends employee_base.html)
    # vs admins (extends admin_dashboard/base.html)
    is_accountant = False
    try:
        employee = request.user.employee_profile
        if employee.role == 'accountant':
            is_accountant = True
    except (Employee.DoesNotExist, AttributeError):
        pass
    
    # Use accountant template if user is accountant, otherwise use admin template
    if is_accountant:
        template = 'crm/accountant_reports.html'
    else:
        template = 'crm/admin_reports.html'
    
    return render(request, template, context)


def _generate_reports_excel(reports, search, date_filter, employee_filter):
    """Generate Excel report for handler reports"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename
    filename_parts = ['EmployeeReports']
    if date_filter:
        filename_parts.append(date_filter.replace('-', ''))
    elif search:
        filename_parts.append('search')
    filename = f"{'_'.join(filename_parts)}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Employee Reports'
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 12
    worksheet.column_dimensions['I'].width = 12
    worksheet.column_dimensions['J'].width = 40
    worksheet.column_dimensions['K'].width = 20
    
    # Header row
    headers = [
        'Employee Name', 'Week Start', 'Week End', 'Weekly Target',
        'Calls Made', 'Successful Calls', 'Success Rate %', 'Target Achievement %',
        'Errands Made', 'Notes', 'Created At'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row = 2
    for report in reports:
        employee_name = report.handler.user.get_full_name() or report.handler.user.username
        success_rate = report.success_rate
        target_achievement = report.target_achievement
        
        values = [
            employee_name,
            report.week_start.strftime('%Y-%m-%d'),
            report.week_end.strftime('%Y-%m-%d'),
            report.weekly_target,
            report.calls_made,
            report.successful_calls,
            f'{success_rate:.2f}%',
            f'{target_achievement:.2f}%',
            report.errands_made,
            report.notes or '',
            report.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col, value in enumerate(values, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [4, 5, 6, 8, 9]:  # Numeric columns
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        row += 1
    
    # Add summary row
    if row > 2:  # Check if we have data rows
        summary_row = row + 1
        worksheet.cell(row=summary_row, column=1, value='SUMMARY').font = Font(bold=True)
        worksheet.cell(row=summary_row, column=4, value=f'=SUM(D2:D{row-1})').font = Font(bold=True)
        worksheet.cell(row=summary_row, column=4).alignment = Alignment(horizontal='right', vertical='center')
        worksheet.cell(row=summary_row, column=5, value=f'=SUM(E2:E{row-1})').font = Font(bold=True)
        worksheet.cell(row=summary_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
        worksheet.cell(row=summary_row, column=6, value=f'=SUM(F2:F{row-1})').font = Font(bold=True)
        worksheet.cell(row=summary_row, column=6).alignment = Alignment(horizontal='right', vertical='center')
        worksheet.cell(row=summary_row, column=9, value=f'=SUM(I2:I{row-1})').font = Font(bold=True)
        worksheet.cell(row=summary_row, column=9).alignment = Alignment(horizontal='right', vertical='center')
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    workbook.save(response)
    return response