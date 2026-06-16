from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Q, Count, Sum, Avg
from django.contrib.auth.models import User
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import json
from io import BytesIO
from django.http import HttpResponse, JsonResponse

# Optional Excel export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except Exception:  # openpyxl not installed in some environments
    Workbook = None

from .models import (
    Asset, AssetModel, AssetCategory, Manufacturer, 
    Location, Department, AssetHistory, MaintenanceRecord
)
from .serializers import (
    AssetSerializer, AssetDetailSerializer, AssetCreateUpdateSerializer,
    AssetModelSerializer, AssetCategorySerializer, ManufacturerSerializer,
    LocationSerializer, DepartmentSerializer, UserSerializer,
    AssetHistorySerializer, MaintenanceRecordSerializer,
    DashboardStatsSerializer, AssetDistributionSerializer,
    AssetTrendSerializer, RecentActivitySerializer
)


class AssetViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing assets
    """
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer
    permission_classes = [AllowAny]  # Temporarily allow for testing
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return AssetDetailSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return AssetCreateUpdateSerializer
        return AssetSerializer
    
    def get_queryset(self):
        queryset = Asset.objects.select_related(
            'model', 'model__manufacturer', 'model__category',
            'assigned_to', 'location', 'department'
        ).prefetch_related(
            'history', 'maintenance_records', 'network_interfaces',
            'software_installations'
        )
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by category
        category_filter = self.request.query_params.get('category')
        if category_filter:
            queryset = queryset.filter(model__category__name__icontains=category_filter)
        
        # Filter by location
        location_filter = self.request.query_params.get('location')
        if location_filter:
            queryset = queryset.filter(location__name__icontains=location_filter)
        
        # Search filter
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(asset_tag__icontains=search) |
                Q(name__icontains=search) |
                Q(model__name__icontains=search) |
                Q(model__manufacturer__name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(assigned_to__username__icontains=search) |
                Q(location__name__icontains=search)
            )
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Get asset history"""
        asset = self.get_object()
        history = asset.history.all()
        serializer = AssetHistorySerializer(history, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def maintenance(self, request, pk=None):
        """Get asset maintenance records"""
        asset = self.get_object()
        maintenance = asset.maintenance_records.all()
        serializer = MaintenanceRecordSerializer(maintenance, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update assets"""
        asset_ids = request.data.get('asset_ids', [])
        update_data = request.data.get('data', {})
        
        if not asset_ids:
            return Response({'error': 'No assets selected'}, status=status.HTTP_400_BAD_REQUEST)
        
        assets = Asset.objects.filter(id__in=asset_ids)
        updated_count = 0
        
        for asset in assets:
            for field, value in update_data.items():
                if hasattr(asset, field):
                    setattr(asset, field, value)
            asset.save()
            updated_count += 1
        
        return Response({'updated': updated_count})
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """Bulk delete assets"""
        asset_ids = request.data.get('asset_ids', [])
        
        if not asset_ids:
            return Response({'error': 'No assets selected'}, status=status.HTTP_400_BAD_REQUEST)
        
        deleted_count = Asset.objects.filter(id__in=asset_ids).delete()[0]
        return Response({'deleted': deleted_count})
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export assets to CSV with improved formatting and alignment"""
        format_type = request.query_params.get('format', 'csv')

        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="assets_export.csv"'

            writer = csv.writer(response, quoting=csv.QUOTE_ALL, escapechar='\\', doublequote=True)

            # Improved column headers with better organization
            headers = [
                'Asset Tag', 'Name', 'Model', 'Manufacturer', 'Category',
                'Serial Number', 'Status', 'Department', 'Assigned To', 'Location',
                'IP Address', 'MAC Address', 'Hostname', 'Device Name',
                'Purchase Date', 'Purchase Cost', 'Warranty Expires',
                'Processor', 'RAM', 'System Type', 'Notes', 'Created Date', 'Last Updated'
            ]
            writer.writerow(headers)

            # Get ALL assets with proper relations, not filtered queryset
            assets = Asset.objects.select_related(
                'model', 'model__manufacturer', 'model__category',
                'assigned_to', 'location', 'department'
            ).prefetch_related(
                'history', 'maintenance_records', 'network_interfaces',
                'software_installations'
            ).all()

            for asset in assets:
                # Format dates properly
                purchase_date = asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else ''
                warranty_expires = asset.warranty_expires.strftime('%Y-%m-%d') if asset.warranty_expires else ''
                created_date = asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else ''
                updated_date = asset.updated_at.strftime('%Y-%m-%d %H:%M:%S') if asset.updated_at else ''

                # Format cost with 2 decimal places
                purchase_cost = f"{asset.purchase_cost:.2f}" if asset.purchase_cost else ''

                # Clean and format notes (remove newlines and limit length for CSV)
                notes = asset.notes.replace('\n', ' ').replace('\r', ' ').strip() if asset.notes else ''

                # Ensure all values are strings and handle None values
                row = [
                    str(asset.asset_tag or ''),
                    str(asset.name or ''),
                    str(asset.model.name if asset.model else ''),
                    str(asset.model.manufacturer.name if asset.model and asset.model.manufacturer else ''),
                    str(asset.model.category.name if asset.model and asset.model.category else ''),
                    str(asset.serial_number or ''),
                    str(asset.get_status_display()),
                    str(asset.department.name if asset.department else ''),
                    str(asset.assigned_to.get_full_name() if asset.assigned_to else ''),
                    str(asset.location.name if asset.location else ''),
                    str(asset.ip_address or ''),
                    str(asset.mac_address or ''),
                    str(asset.hostname or ''),
                    str(asset.device_name or ''),
                    purchase_date,
                    purchase_cost,
                    warranty_expires,
                    str(asset.processor or ''),
                    str(asset.installed_ram or ''),
                    str(asset.system_type or ''),
                    notes,
                    created_date,
                    updated_date
                ]
                writer.writerow(row)

            return response

        # Excel export with proper alignment and column sizing
        if format_type == 'xlsx':
            if not Workbook:
                return Response({'error': 'Excel export not available. Install openpyxl.'}, status=status.HTTP_501_NOT_IMPLEMENTED)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Assets'

            headers = [
                'Asset Tag', 'Name', 'Model', 'Manufacturer', 'Category',
                'Serial Number', 'Status', 'Department', 'Assigned To', 'Location',
                'IP Address', 'MAC Address', 'Hostname', 'Device Name',
                'Purchase Date', 'Purchase Cost', 'Warranty Expires',
                'Processor', 'RAM', 'System Type', 'Notes', 'Created Date', 'Last Updated'
            ]

            # Write header with styling
            header_font = Font(bold=True)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Write data rows
            row_idx = 2
            assets = Asset.objects.select_related(
                'model', 'model__manufacturer', 'model__category',
                'assigned_to', 'location', 'department'
            ).all()

            for asset in assets:
                purchase_date = asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else ''
                warranty_expires = asset.warranty_expires.strftime('%Y-%m-%d') if asset.warranty_expires else ''
                created_date = asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else ''
                updated_date = asset.updated_at.strftime('%Y-%m-%d %H:%M:%S') if asset.updated_at else ''
                purchase_cost = float(asset.purchase_cost) if asset.purchase_cost is not None else ''
                notes = asset.notes.replace('\n', ' ').replace('\r', ' ').strip() if asset.notes else ''

                values = [
                    asset.asset_tag or '',
                    asset.name or '',
                    asset.model.name if asset.model else '',
                    asset.model.manufacturer.name if asset.model and asset.model.manufacturer else '',
                    asset.model.category.name if asset.model and asset.model.category else '',
                    asset.serial_number or '',
                    asset.get_status_display(),
                    asset.department.name if asset.department else '',
                    asset.assigned_to.get_full_name() if asset.assigned_to else '',
                    asset.location.name if asset.location else '',
                    asset.ip_address or '',
                    asset.mac_address or '',
                    asset.hostname or '',
                    asset.device_name or '',
                    purchase_date,
                    purchase_cost,
                    warranty_expires,
                    asset.processor or '',
                    asset.installed_ram or '',
                    asset.system_type or '',
                    notes,
                    created_date,
                    updated_date,
                ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Alignment: left for text, right for numbers/dates
                    if col_idx in (16,):  # Purchase Cost column index
                        cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                row_idx += 1

            # Auto-size columns (basic heuristic)
            for col_idx in range(1, len(headers) + 1):
                column = get_column_letter(col_idx)
                max_len = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        val = str(cell.value) if cell.value is not None else ''
                        max_len = max(max_len, len(val))
                ws.column_dimensions[column].width = min(max(12, max_len + 2), 40)

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Output
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="assets_export.xlsx"'
            return response

        return Response({'error': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)


class AssetModelViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for asset models"""
    queryset = AssetModel.objects.all()
    serializer_class = AssetModelSerializer
    permission_classes = [AllowAny]


class AssetCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for asset categories"""
    queryset = AssetCategory.objects.all()
    serializer_class = AssetCategorySerializer
    permission_classes = [AllowAny]


class ManufacturerViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for manufacturers"""
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [AllowAny]


class LocationViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for locations"""
    queryset = Location.objects.all()
    serializer_class = LocationSerializer
    permission_classes = [AllowAny]


class DepartmentViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for departments"""
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [AllowAny]


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for users"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]


class MaintenanceRecordViewSet(viewsets.ModelViewSet):
    """ViewSet for maintenance records"""
    queryset = MaintenanceRecord.objects.all()
    serializer_class = MaintenanceRecordSerializer
    permission_classes = [AllowAny]
    
    def get_queryset(self):
        queryset = MaintenanceRecord.objects.select_related(
            'asset', 'performed_by'
        )
        
        # Filter by asset
        asset_id = self.request.query_params.get('asset')
        if asset_id:
            queryset = queryset.filter(asset_id=asset_id)
        
        return queryset


# Dashboard API Views
@api_view(['GET'])
@permission_classes([AllowAny])  # Allow health check without authentication
def health_check(request):
    """Health check endpoint"""
    return Response({'status': 'ok', 'timestamp': timezone.now()})


@api_view(['GET'])
@permission_classes([AllowAny])  # Temporarily allow for testing
def dashboard_stats(request):
    """Get dashboard statistics"""
    total_assets = Asset.objects.count()
    active_assets = Asset.objects.filter(status='active').count()
    inactive_assets = Asset.objects.filter(status='inactive').count()
    maintenance_assets = Asset.objects.filter(status='maintenance').count()
    retired_assets = Asset.objects.filter(status='retired').count()
    
    total_value = Asset.objects.aggregate(
        total=Sum('purchase_cost')
    )['total'] or Decimal('0.00')
    
    departments_count = Department.objects.count()
    locations_count = Location.objects.count()
    users_count = User.objects.count()
    
    data = {
        'total_assets': total_assets,
        'active_assets': active_assets,
        'inactive_assets': inactive_assets,
        'maintenance_assets': maintenance_assets,
        'retired_assets': retired_assets,
        'total_value': total_value,
        'departments_count': departments_count,
        'locations_count': locations_count,
        'users_count': users_count
    }
    
    serializer = DashboardStatsSerializer(data)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([AllowAny])  # Temporarily allow for testing
def asset_distribution(request):
    """Get asset distribution by status"""
    total_assets = Asset.objects.count()
    if total_assets == 0:
        return Response([])
    
    status_counts = Asset.objects.values('status').annotate(count=Count('id'))
    
    data = []
    for item in status_counts:
        percentage = (item['count'] / total_assets) * 100
        data.append({
            'status': item['status'],
            'count': item['count'],
            'percentage': round(percentage, 1)
        })
    
    return Response(data)


@api_view(['GET'])
@permission_classes([AllowAny])
def asset_trends(request):
    """Get asset trends over time"""
    months = int(request.query_params.get('months', 6))
    
    # Generate labels for the last N months
    labels = []
    assets_added = []
    maintenance_tasks = []
    
    for i in range(months):
        date = timezone.now() - timedelta(days=30 * i)
        month_start = date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = (month_start + timedelta(days=32)).replace(day=1)
        
        labels.insert(0, date.strftime('%b %Y'))
        
        # Count assets added in this month
        added_count = Asset.objects.filter(
            created_at__gte=month_start,
            created_at__lt=next_month
        ).count()
        assets_added.insert(0, added_count)
        
        # Count maintenance tasks in this month
        maintenance_count = MaintenanceRecord.objects.filter(
            created_at__gte=month_start,
            created_at__lt=next_month
        ).count()
        maintenance_tasks.insert(0, maintenance_count)
    
    data = {
        'labels': labels,
        'assets_added': assets_added,
        'maintenance_tasks': maintenance_tasks
    }
    
    return Response(data)


@api_view(['GET'])
@permission_classes([AllowAny])  # Temporarily allow for testing
def recent_activity(request):
    """Get recent activity"""
    limit = int(request.query_params.get('limit', 10))
    
    # Get recent asset history
    recent_history = AssetHistory.objects.select_related(
        'asset', 'user'
    ).order_by('-timestamp')[:limit]
    
    data = []
    for history in recent_history:
        data.append({
            'action': history.get_action_display(),
            'description': history.description,
            'timestamp': history.timestamp,
            'user': history.user.username if history.user else 'System',
            'asset_tag': history.asset.asset_tag
        })
    
    return Response(data)


@api_view(['GET'])
@permission_classes([AllowAny])
def generate_report(request, report_type):
    """Generate different types of reports"""
    
    if report_type == 'asset-summary':
        return generate_asset_summary_report()
    elif report_type == 'maintenance-schedule':
        return generate_maintenance_schedule_report()
    elif report_type == 'financial-report':
        return generate_financial_report()
    elif report_type == 'assignment-report':
        return generate_assignment_report()
    else:
        return Response({'error': 'Unknown report type'}, status=400)


def generate_asset_summary_report():
    """Generate asset summary report"""
    total_assets = Asset.objects.count()
    active_assets = Asset.objects.filter(status='active').count()
    inactive_assets = Asset.objects.filter(status='inactive').count()
    maintenance_assets = Asset.objects.filter(status='maintenance').count()
    
    # Categories breakdown
    categories = AssetCategory.objects.annotate(
        asset_count=Count('assetmodel__asset')
    ).filter(asset_count__gt=0)
    
    category_data = []
    for category in categories:
        percentage = (category.asset_count / total_assets) * 100 if total_assets > 0 else 0
        category_data.append({
            'name': category.name,
            'count': category.asset_count,
            'percentage': round(percentage, 1)
        })
    
    # Status breakdown
    status_counts = Asset.objects.values('status').annotate(count=Count('id'))
    status_data = []
    for item in status_counts:
        percentage = (item['count'] / total_assets) * 100 if total_assets > 0 else 0
        status_data.append({
            'status': item['status'],
            'name': dict(Asset.STATUS_CHOICES)[item['status']],
            'count': item['count'],
            'percentage': round(percentage, 1)
        })
    
    return Response({
        'total_assets': total_assets,
        'active_assets': active_assets,
        'inactive_assets': inactive_assets,
        'maintenance_assets': maintenance_assets,
        'categories': category_data,
        'statuses': status_data
    })


def generate_maintenance_schedule_report():
    """Generate maintenance schedule report"""
    now = timezone.now()
    
    # Get overdue tasks
    overdue_tasks = MaintenanceRecord.objects.filter(
        scheduled_date__lt=now,
        status__in=['scheduled', 'in_progress']
    )
    
    # Get upcoming tasks (next 30 days)
    upcoming_tasks = MaintenanceRecord.objects.filter(
        scheduled_date__gte=now,
        scheduled_date__lt=now + timedelta(days=30),
        status='scheduled'
    )
    
    # Get all scheduled tasks
    scheduled_tasks = MaintenanceRecord.objects.filter(
        status='scheduled'
    )
    
    # Format tasks for response
    tasks = []
    all_tasks = MaintenanceRecord.objects.filter(
        status__in=['scheduled', 'in_progress']
    ).select_related('asset')
    
    for task in all_tasks:
        tasks.append({
            'asset_tag': task.asset.asset_tag,
            'title': task.title,
            'scheduled_date': task.scheduled_date.isoformat(),
            'status': task.status,
            'priority': 'high' if task.scheduled_date < now else 'medium'
        })
    
    return Response({
        'overdue_count': overdue_tasks.count(),
        'upcoming_count': upcoming_tasks.count(),
        'scheduled_count': scheduled_tasks.count(),
        'tasks': tasks
    })


def generate_financial_report():
    """Generate financial report"""
    total_value = Asset.objects.aggregate(
        total=Sum('purchase_cost')
    )['total'] or Decimal('0.00')
    
    # Calculate depreciated value (simple 20% per year)
    depreciated_value = total_value * Decimal('0.8')  # Simplified calculation
    
    # Calculate maintenance costs
    maintenance_cost = MaintenanceRecord.objects.aggregate(
        total=Sum('cost')
    )['total'] or Decimal('0.00')
    
    # Categories breakdown
    categories = AssetCategory.objects.annotate(
        asset_count=Count('assetmodel__asset'),
        total_value=Sum('assetmodel__asset__purchase_cost'),
        avg_value=Avg('assetmodel__asset__purchase_cost')
    ).filter(asset_count__gt=0)
    
    category_data = []
    for category in categories:
        category_data.append({
            'name': category.name,
            'count': category.asset_count,
            'total_value': category.total_value or 0,
            'average_value': category.avg_value or 0
        })
    
    return Response({
        'total_value': total_value,
        'depreciated_value': depreciated_value,
        'maintenance_cost': maintenance_cost,
        'categories': category_data
    })


def generate_assignment_report():
    """Generate assignment report"""
    assigned_assets = Asset.objects.filter(assigned_to__isnull=False)
    unassigned_assets = Asset.objects.filter(assigned_to__isnull=True)
    active_users = User.objects.filter(assigned_assets__isnull=False).distinct()
    
    # Get assignments by user
    assignments = []
    for user in active_users:
        user_assets = user.assigned_assets.all()
        total_value = user_assets.aggregate(
            total=Sum('purchase_cost')
        )['total'] or Decimal('0.00')
        
        assignments.append({
            'user': user.get_full_name() or user.username,
            'department': user.userprofile.department.name if hasattr(user, 'userprofile') and user.userprofile.department else 'N/A',
            'asset_count': user_assets.count(),
            'total_value': total_value
        })
    
    return Response({
        'assigned_count': assigned_assets.count(),
        'unassigned_count': unassigned_assets.count(),
        'users_count': active_users.count(),
        'assignments': assignments
    })


# Discovery API Views
@api_view(['POST'])
@permission_classes([AllowAny])
def start_network_discovery(request):
    """Start network discovery"""
    # This would integrate with your discovery system
    return Response({'message': 'Network discovery started'})


@api_view(['GET'])
@permission_classes([AllowAny])
def get_discovered_devices(request):
    """Get discovered devices"""
    # This would return discovered devices
    return Response({'devices': []})